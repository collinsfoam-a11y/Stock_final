#!/usr/bin/env python3
"""Import ERP stock data from an XLSX stock export into MongoDB.

The script is intentionally dry-run by default. Use --execute only after
reviewing the preflight summary and approval log entry.
"""

from __future__ import annotations

import argparse
import asyncio
import os
import uuid
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook
from pymongo import ReplaceOne
from pymongo.errors import BulkWriteError


EXPECTED_HEADERS = [
    "Section",
    "Group",
    "Department",
    "Category",
    "Sub Category",
    "Tax Category",
    "Brand",
    "Code",
    "Product",
    "Auto Barcode",
    "CGST %",
    "SGST %",
    "Cost",
    "Std PPrice",
    "Std SPrice",
    "MRP",
    "Stock",
    "Stock Value",
]

HIERARCHY_FIELDS = [
    "section",
    "group",
    "department",
    "category",
    "subcategory",
    "tax_category",
    "brand_name",
]


@dataclass(frozen=True)
class ParsedStock:
    items: list[dict[str, Any]]
    skipped_subtotal_rows: int
    skipped_blank_rows: int
    invalid_rows: list[dict[str, Any]]
    duplicate_item_codes: dict[str, int]
    duplicate_barcodes: dict[str, int]
    bad_barcode_prefix_rows: list[dict[str, Any]]
    negative_stock_rows: list[dict[str, Any]]
    zero_stock_count: int
    stock_total: float
    stock_value_total: float


def utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def clean_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def numeric_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def numeric_or_zero(value: Any) -> float:
    return numeric_or_none(value) or 0.0


def is_total_marker(value: Any) -> bool:
    text = clean_string(value)
    return bool(text and text.lower().endswith("total"))


def is_subtotal_row(values: list[Any]) -> bool:
    return any(is_total_marker(value) for value in values[:9])


def normalize_code(value: Any) -> str | None:
    text = clean_string(value)
    if text is None:
        return None
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return text


def normalize_barcode(value: Any) -> str | None:
    text = clean_string(value)
    if text is None:
        return None
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return text


def read_headers(sheet: Any, header_row: int) -> list[str]:
    row = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    return [str(value).strip() if value is not None else "" for value in row]


def validate_headers(headers: list[str]) -> None:
    expected = EXPECTED_HEADERS
    actual = headers[: len(expected)]
    if actual != expected:
        raise SystemExit(
            "Unexpected XLSX headers.\n"
            f"Expected: {expected}\n"
            f"Actual:   {actual}"
        )


def build_item(
    *,
    row_num: int,
    values: list[Any],
    context: dict[str, str | None],
    import_run_id: str,
    source_path: Path,
    now: datetime,
) -> dict[str, Any]:
    item_code = normalize_code(values[7])
    barcode = normalize_barcode(values[9])
    item_name = clean_string(values[8])

    if item_code is None or barcode is None or item_name is None:
        raise ValueError("missing item code, barcode, or product name")

    stock_qty = numeric_or_zero(values[16])
    cost = numeric_or_none(values[12])
    std_purchase_price = numeric_or_none(values[13])
    std_sales_price = numeric_or_none(values[14])
    mrp = numeric_or_none(values[15])
    cgst = numeric_or_none(values[10])
    sgst = numeric_or_none(values[11])

    item: dict[str, Any] = {
        "item_code": item_code,
        "original_item_code": item_code,
        "item_name": item_name,
        "barcode": barcode,
        "autobarcode": barcode,
        "stock_qty": stock_qty,
        "sql_server_qty": stock_qty,
        "mrp": mrp,
        "cost": cost,
        "standard_purchase_price": std_purchase_price,
        "standard_sales_price": std_sales_price,
        "sale_price": std_sales_price,
        "sales_price": std_sales_price,
        "stock_value": numeric_or_zero(values[17]),
        "cgst_percent": cgst,
        "sgst_percent": sgst,
        "gst_percent": (cgst or 0.0) + (sgst or 0.0),
        "tax_category": context.get("tax_category"),
        "category": context.get("category") or "General",
        "subcategory": context.get("subcategory"),
        "brand_name": context.get("brand_name") or "Default",
        "department": context.get("department"),
        "section": context.get("section"),
        "group": context.get("group"),
        "warehouse": "Main",
        "batch_id": barcode,
        "batch_no": barcode,
        "verified": False,
        "data_complete": False,
        "completion_percentage": 0,
        "missing_fields": ["serial_number"],
        "enrichment_history": [],
        "serial_number": None,
        "condition": None,
        "synced_from_erp": True,
        "synced_from_sql": False,
        "source_import": source_path.name,
        "source_file": str(source_path),
        "source_row": row_num,
        "import_run_id": import_run_id,
        "last_synced": now,
        "synced_at": now,
        "created_at": now,
        "updated_at": now,
    }

    return item


def parse_stock_xlsx(
    path: Path,
    *,
    sheet_name: str | None,
    header_row: int,
    import_run_id: str,
    limit: int | None,
) -> ParsedStock:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        headers = read_headers(sheet, header_row)
        validate_headers(headers)

        now = utc_now()
        context: dict[str, str | None] = {field: None for field in HIERARCHY_FIELDS}
        items: list[dict[str, Any]] = []
        invalid_rows: list[dict[str, Any]] = []
        bad_barcode_prefix_rows: list[dict[str, Any]] = []
        negative_stock_rows: list[dict[str, Any]] = []
        skipped_subtotal_rows = 0
        skipped_blank_rows = 0
        code_counter: Counter[str] = Counter()
        barcode_counter: Counter[str] = Counter()
        stock_total = 0.0
        stock_value_total = 0.0
        zero_stock_count = 0

        for row_num, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            values = list(row[: len(EXPECTED_HEADERS)])
            if all(value in (None, "") for value in values):
                skipped_blank_rows += 1
                continue
            if is_subtotal_row(values):
                skipped_subtotal_rows += 1
                continue

            for idx, field in enumerate(HIERARCHY_FIELDS):
                value = clean_string(values[idx])
                if value is not None:
                    context[field] = value

            if values[7] in (None, "") and values[8] in (None, ""):
                skipped_blank_rows += 1
                continue

            try:
                item = build_item(
                    row_num=row_num,
                    values=values,
                    context=context,
                    import_run_id=import_run_id,
                    source_path=path,
                    now=now,
                )
            except ValueError as exc:
                invalid_rows.append({"row": row_num, "reason": str(exc)})
                continue

            barcode = str(item["barcode"])
            if not (barcode.isdigit() and len(barcode) == 6):
                invalid_rows.append(
                    {"row": row_num, "reason": "barcode must be exactly 6 digits"}
                )
                continue
            if barcode[:2] not in {"51", "52", "53"}:
                bad_barcode_prefix_rows.append({"row": row_num, "barcode": barcode})

            stock_qty = float(item["stock_qty"])
            if stock_qty < 0:
                negative_stock_rows.append(
                    {
                        "row": row_num,
                        "item_code": item["item_code"],
                        "barcode": barcode,
                        "stock_qty": stock_qty,
                    }
                )
            if abs(stock_qty) < 0.000001:
                zero_stock_count += 1

            items.append(item)
            code_counter[str(item["item_code"])] += 1
            barcode_counter[barcode] += 1
            stock_total += stock_qty
            stock_value_total += float(item["stock_value"])

            if limit is not None and len(items) >= limit:
                break

        duplicate_item_codes = {code: count for code, count in code_counter.items() if count > 1}
        duplicate_barcodes = {
            barcode: count for barcode, count in barcode_counter.items() if count > 1
        }

        return ParsedStock(
            items=items,
            skipped_subtotal_rows=skipped_subtotal_rows,
            skipped_blank_rows=skipped_blank_rows,
            invalid_rows=invalid_rows,
            duplicate_item_codes=duplicate_item_codes,
            duplicate_barcodes=duplicate_barcodes,
            bad_barcode_prefix_rows=bad_barcode_prefix_rows,
            negative_stock_rows=negative_stock_rows,
            zero_stock_count=zero_stock_count,
            stock_total=stock_total,
            stock_value_total=stock_value_total,
        )
    finally:
        workbook.close()


async def get_unique_item_code_index(db: Any) -> str | None:
    indexes = await db.erp_items.list_indexes().to_list(length=100)
    for index in indexes:
        key = dict(index.get("key", {}))
        if key == {"item_code": 1} and index.get("unique") is True:
            return str(index["name"])
    return None


async def explain_existing_state(db: Any, items: list[dict[str, Any]]) -> dict[str, Any]:
    barcodes = [item["barcode"] for item in items]
    item_codes = [item["item_code"] for item in items]

    existing_count = await db.erp_items.count_documents({})
    existing_by_barcode = await db.erp_items.count_documents({"barcode": {"$in": barcodes}})
    existing_by_item_code = await db.erp_items.count_documents({"item_code": {"$in": item_codes}})
    unique_item_code_index = await get_unique_item_code_index(db)

    return {
        "existing_count": existing_count,
        "existing_by_barcode": existing_by_barcode,
        "existing_by_item_code": existing_by_item_code,
        "unique_item_code_index": unique_item_code_index,
    }


def print_summary(parsed: ParsedStock, db_state: dict[str, Any], *, execute: bool) -> None:
    print("mode=", "execute" if execute else "dry-run")
    print("parsed_product_rows=", len(parsed.items))
    print("skipped_subtotal_rows=", parsed.skipped_subtotal_rows)
    print("skipped_blank_rows=", parsed.skipped_blank_rows)
    print("invalid_rows=", len(parsed.invalid_rows))
    print("duplicate_item_code_keys=", len(parsed.duplicate_item_codes))
    print("duplicate_item_code_rows=", sum(parsed.duplicate_item_codes.values()))
    print("duplicate_barcode_keys=", len(parsed.duplicate_barcodes))
    print("bad_barcode_prefix_rows=", len(parsed.bad_barcode_prefix_rows))
    print("negative_stock_rows=", len(parsed.negative_stock_rows))
    print("zero_stock_rows=", parsed.zero_stock_count)
    print("stock_total=", round(parsed.stock_total, 6))
    print("stock_value_total=", round(parsed.stock_value_total, 6))
    print("mongo_existing_erp_items=", db_state["existing_count"])
    print("mongo_existing_by_barcode=", db_state["existing_by_barcode"])
    print("mongo_existing_by_item_code=", db_state["existing_by_item_code"])
    print("mongo_unique_item_code_index=", db_state["unique_item_code_index"])

    if parsed.duplicate_item_codes:
        samples = list(parsed.duplicate_item_codes.items())[:10]
        print("duplicate_item_code_samples=", samples)
    if parsed.negative_stock_rows:
        print("negative_stock_samples=", parsed.negative_stock_rows[:10])
    if parsed.invalid_rows:
        print("invalid_row_samples=", parsed.invalid_rows[:10])


async def ensure_item_code_index_allows_duplicates(db: Any, *, drop_unique: bool) -> None:
    unique_index_name = await get_unique_item_code_index(db)
    if not unique_index_name:
        return
    if not drop_unique:
        raise SystemExit(
            "MongoDB has a unique item_code index, but the XLSX has duplicate item codes. "
            "Re-run with --drop-unique-item-code-index after approval, or import will fail."
        )

    await db.erp_items.drop_index(unique_index_name)
    await db.erp_items.create_index([("item_code", 1)], name=unique_index_name)


async def execute_import(
    db: Any,
    items: list[dict[str, Any]],
    *,
    batch_size: int,
    allow_existing: bool,
) -> dict[str, Any]:
    existing_count = await db.erp_items.count_documents({})
    if existing_count and not allow_existing:
        raise SystemExit(
            f"erp_items already has {existing_count} documents. "
            "Use --allow-existing to upsert by barcode."
        )

    inserted_or_updated = 0
    errors: list[dict[str, Any]] = []
    for start in range(0, len(items), batch_size):
        batch = items[start : start + batch_size]
        operations = [
            ReplaceOne({"barcode": item["barcode"]}, item, upsert=True) for item in batch
        ]
        try:
            result = await db.erp_items.bulk_write(operations, ordered=False)
            inserted_or_updated += (
                result.upserted_count + result.modified_count + result.matched_count
            )
        except BulkWriteError as exc:
            errors.extend(exc.details.get("writeErrors", []))

    return {
        "attempted": len(items),
        "inserted_or_updated": inserted_or_updated,
        "errors": len(errors),
        "error_samples": errors[:5],
    }


async def async_main(args: argparse.Namespace) -> int:
    source_path = Path(args.xlsx).expanduser().resolve()
    if not source_path.exists():
        raise SystemExit(f"XLSX file not found: {source_path}")

    import_run_id = args.import_run_id or f"stock_xlsx_{uuid.uuid4().hex[:12]}"
    parsed = parse_stock_xlsx(
        source_path,
        sheet_name=args.sheet,
        header_row=args.header_row,
        import_run_id=import_run_id,
        limit=args.limit,
    )

    client = AsyncIOMotorClient(args.mongo_url, serverSelectionTimeoutMS=5000)
    try:
        await client.admin.command("ping")
        db = client[args.db_name]
        db_state = await explain_existing_state(db, parsed.items)
        print("import_run_id=", import_run_id)
        print("mongo_url=", args.mongo_url)
        print("db_name=", args.db_name)
        print_summary(parsed, db_state, execute=args.execute)

        if not args.execute:
            print("No data was written. Re-run with --execute after approval.")
            return 0

        if parsed.invalid_rows:
            raise SystemExit("Refusing import because invalid rows were found.")
        if parsed.duplicate_barcodes:
            raise SystemExit("Refusing import because duplicate barcodes were found.")
        if parsed.duplicate_item_codes:
            await ensure_item_code_index_allows_duplicates(
                db, drop_unique=args.drop_unique_item_code_index
            )

        result = await execute_import(
            db,
            parsed.items,
            batch_size=args.batch_size,
            allow_existing=args.allow_existing,
        )
        print("write_result=", result)
        final_count = await db.erp_items.count_documents({"import_run_id": import_run_id})
        print("import_run_id_document_count=", final_count)
        return 0 if result["errors"] == 0 else 1
    finally:
        client.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", help="Path to STOCK.xlsx or compatible stock export.")
    parser.add_argument("--sheet", default=None, help="Worksheet name. Defaults to first sheet.")
    parser.add_argument("--header-row", type=int, default=2)
    parser.add_argument("--mongo-url", default=os.getenv("MONGO_URL", "mongodb://localhost:27017"))
    parser.add_argument("--db-name", default=os.getenv("DB_NAME", "stock_verification"))
    parser.add_argument("--batch-size", type=int, default=500)
    parser.add_argument("--limit", type=int, default=None, help="Parse/import only N product rows.")
    parser.add_argument("--import-run-id", default=None)
    parser.add_argument("--allow-existing", action="store_true")
    parser.add_argument(
        "--drop-unique-item-code-index",
        action="store_true",
        help="Drop and recreate item_code index as non-unique before importing duplicates.",
    )
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Write to MongoDB. Without this flag the script only prints a dry-run summary.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return asyncio.run(async_main(args))


if __name__ == "__main__":
    raise SystemExit(main())
