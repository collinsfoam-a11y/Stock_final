"""
Tests for BSR ERPNext export file generation: CSV downloads generated
exclusively from an APPROVED, hash-verified erpnext_export_previews
snapshot. Never re-reads count_lines/erp_items, never recalculates
quantities, never pushes to ERPNext.
"""

import csv
import io
from datetime import datetime, timezone
from unittest.mock import patch

import pytest

from backend.api.erpnext_exports_api import download_erpnext_export_file
from backend.services.erpnext_export_file_service import (
    ErpNextExportFileError,
    ErpNextExportFileIntegrityError,
    ErpNextExportFileNotFoundError,
    ErpNextExportFileService,
)
from backend.services.erpnext_export_service import ErpNextExportService
from backend.tests.utils.in_memory_db import InMemoryDatabase

CURRENT_USER = {"username": "supervisor1", "role": "supervisor"}
ADMIN_USER = {"username": "admin1", "role": "admin"}
DEFAULT_WAREHOUSE = "WH-1"
DEFAULT_UOM = "PCS"
DEFAULT_COMPANY = "Default Co"


async def _seed_finalized_session(db: InMemoryDatabase, session_id: str) -> None:
    await db.sessions.insert_one(
        {
            "id": session_id,
            "session_id": session_id,
            "status": "COMPLETED",
            "warehouse": DEFAULT_WAREHOUSE,
            "finalized_at": datetime.now(timezone.utc),
            "version": 1,
        }
    )


async def _seed_default_mappings(db: InMemoryDatabase) -> None:
    await db.erpnext_warehouse_mappings.insert_one(
        {
            "mapping_id": "map-wh-1",
            "stock_verify_warehouse_id": DEFAULT_WAREHOUSE,
            "stock_verify_warehouse_name": DEFAULT_WAREHOUSE,
            "erpnext_warehouse": "Stores - CO",
            "company": DEFAULT_COMPANY,
            "is_active": True,
            "created_by": "admin1",
            "created_at": datetime.now(timezone.utc),
            "updated_by": None,
            "updated_at": None,
        }
    )
    await db.erpnext_uom_mappings.insert_one(
        {
            "mapping_id": "map-uom-1",
            "stock_verify_uom": DEFAULT_UOM,
            "erpnext_uom": "Nos",
            "conversion_factor": 1.0,
            "is_active": True,
            "created_by": "admin1",
            "created_at": datetime.now(timezone.utc),
            "updated_by": None,
            "updated_at": None,
        }
    )


async def _seed_erp_item(db: InMemoryDatabase, item_code: str, **overrides) -> None:
    doc = {
        "item_code": item_code,
        "item_name": f"Item {item_code}",
        "warehouse": DEFAULT_WAREHOUSE,
        "uom_code": DEFAULT_UOM,
        "stock_qty": 90.0,
        "mrp": 25.0,
        "last_synced": datetime.now(timezone.utc),
    }
    doc.update(overrides)
    await db.erp_items.insert_one(doc)


async def _seed_approved_line(db: InMemoryDatabase, session_id: str, item_code: str, **overrides):
    doc = {
        "id": f"line-{item_code}-{session_id}",
        "session_id": session_id,
        "item_code": item_code,
        "item_name": f"Item {item_code}",
        "counted_qty": 100.0,
        "erp_qty": 100.0,
        "approval_status": "APPROVED",
        "status": "approved",
        "uom_code": DEFAULT_UOM,
    }
    doc.update(overrides)
    await db.count_lines.insert_one(doc)
    return doc


async def _approve_ready_preview(
    db: InMemoryDatabase, session_id: str, item_code: str, *, line_overrides=None
) -> dict:
    """Seeds a fully-mapped, approvable session/item/line, generates the
    preview, approves it, and returns the stored (approved) document."""
    await _seed_finalized_session(db, session_id)
    await _seed_default_mappings(db)
    await _seed_erp_item(db, item_code)
    await _seed_approved_line(db, session_id, item_code, **(line_overrides or {}))
    preview = await ErpNextExportService(db).generate_preview(
        session_id=session_id, mode="STOCK_ADJUSTMENT", company=DEFAULT_COMPANY, current_user=CURRENT_USER
    )
    await ErpNextExportService(db).approve_preview(
        export_id=preview["export_id"], current_user=ADMIN_USER
    )
    return await db.erpnext_export_previews.find_one({"export_id": preview["export_id"]})


def _parse_csv(content: bytes) -> list[dict]:
    return list(csv.DictReader(io.StringIO(content.decode("utf-8"))))


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    rows = [dict(zip(headers, row)) for row in rows_iter]
    return headers, rows


# ==================== Source-of-truth guards ====================


@pytest.mark.asyncio
async def test_cannot_download_file_for_non_approved_preview():
    db = InMemoryDatabase()
    await _seed_finalized_session(db, "sess-1")
    await _seed_default_mappings(db)
    await _seed_erp_item(db, "ITEM-1")
    await _seed_approved_line(db, "sess-1", "ITEM-1")
    preview = await ErpNextExportService(db).generate_preview(
        session_id="sess-1", mode="STOCK_ADJUSTMENT", company=DEFAULT_COMPANY, current_user=CURRENT_USER
    )
    assert preview["status"] == "READY_FOR_APPROVAL"  # not approved yet

    service = ErpNextExportFileService(db)
    with pytest.raises(ErpNextExportFileError):
        await service.generate_file(
            export_id=preview["export_id"], file_type="stock_entry", current_user=ADMIN_USER
        )


@pytest.mark.asyncio
async def test_cannot_download_file_if_approval_hash_missing():
    """Simulates a malformed/legacy APPROVED record with no approval_hash
    recorded -- must refuse rather than trust it."""
    db = InMemoryDatabase()
    await db.erpnext_export_previews.insert_one(
        {
            "export_id": "export-no-hash",
            "export_version": 1,
            "session_id": "sess-no-hash",
            "mode": "STOCK_ADJUSTMENT",
            "company": DEFAULT_COMPANY,
            "status": "APPROVED",
            "approval_hash": None,
            "summary": {},
            "rows": [{"item_code": "ITEM-X"}],
            "blockers": [],
            "warnings": [],
        }
    )
    service = ErpNextExportFileService(db)

    with pytest.raises(ErpNextExportFileError):
        await service.generate_file(
            export_id="export-no-hash", file_type="stock_entry", current_user=ADMIN_USER
        )


@pytest.mark.asyncio
async def test_cannot_download_missing_preview():
    db = InMemoryDatabase()
    service = ErpNextExportFileService(db)

    with pytest.raises(ErpNextExportFileNotFoundError):
        await service.generate_file(
            export_id="does-not-exist", file_type="stock_entry", current_user=ADMIN_USER
        )


@pytest.mark.asyncio
async def test_hash_drift_from_recorded_approval_hash_is_a_critical_error():
    """If the frozen payload no longer matches its recorded approval_hash
    (e.g. tampered document), file generation must refuse, not silently
    export whatever is there."""
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-tamper", "ITEM-TAMPER")
    tampered_rows = approved["rows"]
    tampered_rows[0]["counted_qty"] = 999999.0  # tamper with a frozen row post-approval
    await db.erpnext_export_previews.update_one(
        {"export_id": approved["export_id"]},
        {"$set": {"rows": tampered_rows}},
    )
    service = ErpNextExportFileService(db)

    with pytest.raises(ErpNextExportFileIntegrityError):
        await service.generate_file(
            export_id=approved["export_id"], file_type="stock_entry", current_user=ADMIN_USER
        )


# ==================== Formula / column safety ====================


@pytest.mark.asyncio
async def test_stock_entry_csv_uses_opening_qty_not_variance_or_adjustment():
    db = InMemoryDatabase()
    # baseline(erp_qty)=100, existing_sql_qty(stock_qty)=90, counted_qty=95
    # -> variance = 95-100 = -5, adjustment = 95-90 = 5, opening = 95
    approved = await _approve_ready_preview(
        db, "sess-2", "ITEM-2", line_overrides={"counted_qty": 95.0, "erp_qty": 100.0}
    )
    service = ErpNextExportFileService(db)

    result = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_entry", current_user=ADMIN_USER
    )
    rows = _parse_csv(result["content"])

    assert len(rows) == 1
    assert float(rows[0]["Qty"]) == 95.0  # erpnext_opening_qty, i.e. counted_qty
    assert rows[0]["Item Code"] == "ITEM-2"
    assert rows[0]["Target Warehouse"] == "Stores - CO"
    assert rows[0]["Company"] == DEFAULT_COMPANY
    assert rows[0]["Is Opening"] == "Yes"
    assert rows[0]["Stock Entry Type"] == "Material Receipt"


@pytest.mark.asyncio
async def test_stock_reconciliation_csv_uses_counted_qty_and_adjustment_qty():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(
        db, "sess-3", "ITEM-3", line_overrides={"counted_qty": 95.0, "erp_qty": 100.0}
    )
    service = ErpNextExportFileService(db)

    result = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_reconciliation", current_user=ADMIN_USER
    )
    rows = _parse_csv(result["content"])

    assert len(rows) == 1
    row = rows[0]
    assert float(row["Qty"]) == 95.0  # counted_qty
    assert float(row["Current Qty"]) == 90.0  # existing_sql_qty (erp_item.stock_qty)
    assert float(row["Difference Qty"]) == 5.0  # erpnext_adjustment_qty = counted - existing_sql
    # Never the stock_verify_variance (95-100=-5); must be the adjustment (5).
    assert float(row["Difference Qty"]) != 95.0 - 100.0


@pytest.mark.asyncio
async def test_uom_conversion_factor_exports_in_stock_entry_csv_and_xlsx():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-uom-1", "ITEM-UOM-1")
    service = ErpNextExportFileService(db)

    csv_result = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_entry", file_format="csv", current_user=ADMIN_USER
    )
    csv_rows = _parse_csv(csv_result["content"])
    xlsx_result = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_entry", file_format="xlsx", current_user=ADMIN_USER
    )
    xlsx_headers, xlsx_rows = _parse_xlsx(xlsx_result["content"])

    assert "UOM Conversion Factor" in csv_rows[0]
    assert float(csv_rows[0]["UOM Conversion Factor"]) == 1.0
    assert "UOM Conversion Factor" in xlsx_headers
    assert xlsx_rows[0]["UOM Conversion Factor"] == 1.0


@pytest.mark.asyncio
async def test_uom_conversion_factor_exports_in_stock_reconciliation_csv_and_xlsx():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-uom-2", "ITEM-UOM-2")
    service = ErpNextExportFileService(db)

    csv_result = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_reconciliation", file_format="csv", current_user=ADMIN_USER
    )
    csv_rows = _parse_csv(csv_result["content"])
    xlsx_result = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_reconciliation", file_format="xlsx", current_user=ADMIN_USER
    )
    xlsx_headers, xlsx_rows = _parse_xlsx(xlsx_result["content"])

    assert "UOM Conversion Factor" in csv_rows[0]
    assert float(csv_rows[0]["UOM Conversion Factor"]) == 1.0
    assert "UOM Conversion Factor" in xlsx_headers
    assert xlsx_rows[0]["UOM Conversion Factor"] == 1.0


@pytest.mark.asyncio
async def test_serials_csv_creates_one_row_per_serial():
    db = InMemoryDatabase()
    await _seed_finalized_session(db, "sess-4")
    await _seed_default_mappings(db)
    await _seed_erp_item(db, "ITEM-4", is_serialized=True)
    await _seed_approved_line(
        db,
        "sess-4",
        "ITEM-4",
        counted_qty=2.0,
        erp_qty=2.0,
        serial_numbers=["SN-A", "SN-B"],
    )
    preview = await ErpNextExportService(db).generate_preview(
        session_id="sess-4", mode="STOCK_ADJUSTMENT", company=DEFAULT_COMPANY, current_user=CURRENT_USER
    )
    assert preview["status"] == "READY_FOR_APPROVAL"
    await ErpNextExportService(db).approve_preview(
        export_id=preview["export_id"], current_user=ADMIN_USER
    )

    service = ErpNextExportFileService(db)
    result = await service.generate_file(
        export_id=preview["export_id"], file_type="serials", current_user=ADMIN_USER
    )
    rows = _parse_csv(result["content"])

    assert len(rows) == 2
    assert {r["Serial No"] for r in rows} == {"SN-A", "SN-B"}
    assert all(r["Item Code"] == "ITEM-4" for r in rows)
    assert all(r["Status"] == "Active" for r in rows)


@pytest.mark.asyncio
async def test_batches_csv_creates_unique_batch_rows_only():
    db = InMemoryDatabase()
    await _seed_finalized_session(db, "sess-5")
    await _seed_default_mappings(db)
    await _seed_erp_item(db, "ITEM-5A")
    await _seed_erp_item(db, "ITEM-5B")
    await _seed_approved_line(
        db, "sess-5", "ITEM-5A", batch_id="BATCH-1", manufacturing_date="2026-01-01"
    )
    # Second line, different item, SAME batch_id -- must collapse to one row.
    await _seed_approved_line(
        db,
        "sess-5",
        "ITEM-5B",
        batch_id="BATCH-1",
        manufacturing_date="2026-02-02",
        id="line-ITEM-5B-sess-5-dup",
    )
    preview = await ErpNextExportService(db).generate_preview(
        session_id="sess-5", mode="STOCK_ADJUSTMENT", company=DEFAULT_COMPANY, current_user=CURRENT_USER
    )
    await ErpNextExportService(db).approve_preview(
        export_id=preview["export_id"], current_user=ADMIN_USER
    )

    service = ErpNextExportFileService(db)
    result = await service.generate_file(
        export_id=preview["export_id"], file_type="batches", current_user=ADMIN_USER
    )
    rows = _parse_csv(result["content"])

    assert len(rows) == 1
    assert rows[0]["Batch ID"] == "BATCH-1"
    assert rows[0]["Manufacturing Date"] == "2026-01-01"  # first occurrence wins


@pytest.mark.asyncio
async def test_photo_manifest_exports_references_not_blob():
    from backend.services.erpnext_export_photo_manifest_service import (
        ErpNextExportPhotoManifestService,
    )

    db = InMemoryDatabase()
    await _seed_finalized_session(db, "sess-6")
    await _seed_default_mappings(db)
    await _seed_erp_item(db, "ITEM-6")
    await _seed_approved_line(
        db,
        "sess-6",
        "ITEM-6",
        photo_proofs=[
            {"id": "photo-1", "url": "https://example.com/photo1.jpg", "timestamp": "2026-01-01T00:00:00Z"}
        ],
        photo_base64="this-should-never-appear-in-any-csv-output==",
    )
    preview = await ErpNextExportService(db).generate_preview(
        session_id="sess-6", mode="STOCK_ADJUSTMENT", company=DEFAULT_COMPANY, current_user=CURRENT_USER
    )
    await ErpNextExportService(db).approve_preview(
        export_id=preview["export_id"], current_user=ADMIN_USER
    )
    # Photo-manifest export is sourced exclusively from the manifest
    # snapshot (Part I) -- must be created first.
    await ErpNextExportPhotoManifestService(db).create_manifest(
        export_id=preview["export_id"], current_user=ADMIN_USER
    )

    service = ErpNextExportFileService(db)
    result = await service.generate_file(
        export_id=preview["export_id"], file_type="photo_manifest", current_user=ADMIN_USER
    )
    rows = _parse_csv(result["content"])

    # Both the EXTERNAL_REFERENCE_ONLY (photo_proofs) entry and the DURABLE
    # (photo_base64) entry must be present -- not just the URL reference.
    assert len(rows) == 2
    url_row = next(r for r in rows if r["Durability Status"] == "EXTERNAL_REFERENCE_ONLY")
    assert url_row["Photo ID"] == "photo-1"
    assert url_row["Public Reference"] == "https://example.com/photo1.jpg"
    durable_row = next(r for r in rows if r["Durability Status"] == "DURABLE")
    assert durable_row["Content Hash"]
    assert "this-should-never-appear-in-any-csv-output==" not in result["content"].decode("utf-8")


@pytest.mark.asyncio
async def test_durable_inline_photo_entries_export_in_xlsx():
    from backend.services.erpnext_export_photo_manifest_service import (
        ErpNextExportPhotoManifestService,
    )

    db = InMemoryDatabase()
    await _seed_finalized_session(db, "sess-6b")
    await _seed_default_mappings(db)
    await _seed_erp_item(db, "ITEM-6B")
    await _seed_approved_line(
        db, "sess-6b", "ITEM-6B", photo_base64="aW5saW5lLXBob3RvLWJ5dGVz"
    )
    preview = await ErpNextExportService(db).generate_preview(
        session_id="sess-6b", mode="STOCK_ADJUSTMENT", company=DEFAULT_COMPANY, current_user=CURRENT_USER
    )
    await ErpNextExportService(db).approve_preview(export_id=preview["export_id"], current_user=ADMIN_USER)
    await ErpNextExportPhotoManifestService(db).create_manifest(
        export_id=preview["export_id"], current_user=ADMIN_USER
    )

    service = ErpNextExportFileService(db)
    result = await service.generate_file(
        export_id=preview["export_id"], file_type="photo_manifest", file_format="xlsx", current_user=ADMIN_USER
    )
    headers, rows = _parse_xlsx(result["content"])

    assert len(rows) == 1
    assert rows[0]["Durability Status"] == "DURABLE"
    assert rows[0]["Content Hash"]
    assert "aW5saW5lLXBob3RvLWJ5dGVz" not in str(result["content"])


@pytest.mark.asyncio
async def test_empty_serial_batch_photo_csv_returns_headers_only():
    db = InMemoryDatabase()
    # No serials, no batch_id, no photo_proofs anywhere.
    approved = await _approve_ready_preview(db, "sess-7", "ITEM-7")
    service = ErpNextExportFileService(db)

    for file_type, expected_header in (
        ("serials", "Serial No"),
        ("batches", "Batch ID"),
        ("photo_manifest", "Photo ID"),
    ):
        result = await service.generate_file(
            export_id=approved["export_id"], file_type=file_type, current_user=ADMIN_USER
        )
        rows = _parse_csv(result["content"])
        assert rows == []
        assert expected_header in result["content"].decode("utf-8").splitlines()[0]


# ==================== Hashing / idempotency / audit ====================


@pytest.mark.asyncio
async def test_file_hash_stored_and_matches_returned_content():
    import hashlib

    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-8", "ITEM-8")
    service = ErpNextExportFileService(db)

    result = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_entry", current_user=ADMIN_USER
    )

    assert result["file_hash"] == hashlib.sha256(result["content"]).hexdigest()
    stored = await db.erpnext_export_previews.find_one({"export_id": approved["export_id"]})
    # Hashes are keyed by file_type AND format (e.g. "stock_entry:csv") so
    # the same file type in a different format is tracked independently.
    assert stored["file_hashes"]["stock_entry:csv"] == result["file_hash"]
    assert stored["file_generated_by"]["stock_entry:csv"] == "admin1"
    assert stored["file_generated_at"]["stock_entry:csv"] is not None


@pytest.mark.asyncio
async def test_repeated_download_with_same_preview_is_idempotent():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-9", "ITEM-9")
    service = ErpNextExportFileService(db)

    first = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_entry", current_user=ADMIN_USER
    )
    second = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_entry", current_user=ADMIN_USER
    )

    assert first["content"] == second["content"]
    assert first["file_hash"] == second["file_hash"]
    # Only the first generation should have written an audit event.
    audit_entries = await db.audit_logs.find(
        {"entity_id": approved["export_id"], "details.file_type": "stock_entry"}
    ).to_list(length=10)
    assert len(audit_entries) == 1


@pytest.mark.asyncio
async def test_file_generation_writes_audit_event():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-10", "ITEM-10")
    service = ErpNextExportFileService(db)

    result = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_reconciliation", current_user=ADMIN_USER
    )

    entry = await db.audit_logs.find_one(
        {"entity_id": approved["export_id"], "event_type": "EXPORT_FILE_GENERATED"}
    )
    assert entry is not None
    assert entry["details"]["file_type"] == "stock_reconciliation"
    assert entry["details"]["file_hash"] == result["file_hash"]
    assert entry["details"]["company"] == DEFAULT_COMPANY
    assert entry["details"]["line_count"] == 1


@pytest.mark.asyncio
async def test_file_generator_does_not_read_or_mutate_count_lines():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-11", "ITEM-11")
    before = await db.count_lines.find({"session_id": "sess-11"}).to_list(length=10)

    with patch.object(
        db.count_lines, "find", side_effect=AssertionError("must not read count_lines")
    ):
        service = ErpNextExportFileService(db)
        await service.generate_file(
            export_id=approved["export_id"], file_type="stock_entry", current_user=ADMIN_USER
        )

    after = await db.count_lines.find({"session_id": "sess-11"}).to_list(length=10)
    assert before == after


@pytest.mark.asyncio
async def test_approved_preview_rows_remain_unchanged_after_download():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-12", "ITEM-12")
    rows_before = approved["rows"]
    approval_hash_before = approved["approval_hash"]

    service = ErpNextExportFileService(db)
    for file_type in ("stock_entry", "stock_reconciliation", "serials", "batches", "photo_manifest"):
        await service.generate_file(
            export_id=approved["export_id"], file_type=file_type, current_user=ADMIN_USER
        )

    stored = await db.erpnext_export_previews.find_one({"export_id": approved["export_id"]})
    assert stored["rows"] == rows_before
    assert stored["approval_hash"] == approval_hash_before
    assert stored["status"] == "APPROVED"


@pytest.mark.asyncio
async def test_router_download_endpoint_maps_errors_and_returns_csv(monkeypatch):
    db = InMemoryDatabase()
    monkeypatch.setattr("backend.api.erpnext_exports_api.get_db", lambda: db)

    from fastapi import HTTPException

    with pytest.raises(HTTPException) as exc_info:
        await download_erpnext_export_file(
            "does-not-exist", "stock-entry.csv", current_user=CURRENT_USER
        )
    assert exc_info.value.status_code == 404

    with pytest.raises(HTTPException) as exc_info:
        await download_erpnext_export_file(
            "anything", "unknown-file.csv", current_user=CURRENT_USER
        )
    assert exc_info.value.status_code == 404

    approved = await _approve_ready_preview(db, "sess-13", "ITEM-13")
    response = await download_erpnext_export_file(
        approved["export_id"], "stock-entry.csv", current_user=CURRENT_USER
    )
    assert response.media_type == "text/csv"
    assert b"ITEM-13" in response.body


# ==================== XLSX generation ====================


@pytest.mark.asyncio
async def test_xlsx_generation_requires_approved_preview():
    db = InMemoryDatabase()
    await _seed_finalized_session(db, "sess-14")
    await _seed_default_mappings(db)
    await _seed_erp_item(db, "ITEM-14")
    await _seed_approved_line(db, "sess-14", "ITEM-14")
    preview = await ErpNextExportService(db).generate_preview(
        session_id="sess-14", mode="STOCK_ADJUSTMENT", company=DEFAULT_COMPANY, current_user=CURRENT_USER
    )
    assert preview["status"] == "READY_FOR_APPROVAL"

    service = ErpNextExportFileService(db)
    with pytest.raises(ErpNextExportFileError):
        await service.generate_file(
            export_id=preview["export_id"],
            file_type="stock_entry",
            file_format="xlsx",
            current_user=ADMIN_USER,
        )


@pytest.mark.asyncio
async def test_xlsx_generation_blocks_pending_correction_proposals():
    from backend.services.erpnext_export_correction_service import ErpNextExportCorrectionService

    db = InMemoryDatabase()
    await _seed_finalized_session(db, "sess-15")
    await _seed_default_mappings(db)
    await _seed_erp_item(db, "ITEM-15")
    await _seed_approved_line(db, "sess-15", "ITEM-15")
    preview = await ErpNextExportService(db).generate_preview(
        session_id="sess-15", mode="STOCK_ADJUSTMENT", company=DEFAULT_COMPANY, current_user=CURRENT_USER
    )
    row_key = preview["rows"][0]["count_line_id"]
    await ErpNextExportCorrectionService(db).create_proposal(
        export_id=preview["export_id"],
        row_key=row_key,
        proposed_changes={"batch_no": "BATCH-X"},
        reason="Needs review",
        current_user=CURRENT_USER,
    )
    # Bypass approve_preview's own pending-proposal guard to isolate the
    # file service's independent check.
    await db.erpnext_export_previews.update_one(
        {"export_id": preview["export_id"]},
        {
            "$set": {
                "status": "APPROVED",
                "approved_by": "admin1",
                "approved_at": datetime.now(timezone.utc),
                "approval_hash": ErpNextExportService.compute_approval_hash(preview),
            }
        },
    )

    service = ErpNextExportFileService(db)
    with pytest.raises(ErpNextExportFileError):
        await service.generate_file(
            export_id=preview["export_id"],
            file_type="stock_entry",
            file_format="xlsx",
            current_user=ADMIN_USER,
        )


@pytest.mark.asyncio
async def test_xlsx_generation_verifies_approval_hash():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-16", "ITEM-16")
    tampered_rows = approved["rows"]
    tampered_rows[0]["counted_qty"] = 999999.0
    await db.erpnext_export_previews.update_one(
        {"export_id": approved["export_id"]}, {"$set": {"rows": tampered_rows}}
    )

    service = ErpNextExportFileService(db)
    with pytest.raises(ErpNextExportFileIntegrityError):
        await service.generate_file(
            export_id=approved["export_id"],
            file_type="stock_entry",
            file_format="xlsx",
            current_user=ADMIN_USER,
        )


@pytest.mark.asyncio
async def test_xlsx_content_uses_approved_preview_rows_only():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(
        db, "sess-17", "ITEM-17", line_overrides={"counted_qty": 95.0, "erp_qty": 100.0}
    )
    service = ErpNextExportFileService(db)

    result = await service.generate_file(
        export_id=approved["export_id"],
        file_type="stock_entry",
        file_format="xlsx",
        current_user=ADMIN_USER,
    )
    headers, rows = _parse_xlsx(result["content"])

    assert len(rows) == 1
    assert rows[0]["Item Code"] == "ITEM-17"
    assert rows[0]["Qty"] == 95.0  # erpnext_opening_qty, from the frozen row


@pytest.mark.asyncio
async def test_xlsx_headers_match_csv_headers():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-18", "ITEM-18")
    service = ErpNextExportFileService(db)

    for file_type in ("stock_entry", "stock_reconciliation", "serials", "batches"):
        csv_result = await service.generate_file(
            export_id=approved["export_id"],
            file_type=file_type,
            file_format="csv",
            current_user=ADMIN_USER,
        )
        xlsx_result = await service.generate_file(
            export_id=approved["export_id"],
            file_type=file_type,
            file_format="xlsx",
            current_user=ADMIN_USER,
        )
        csv_rows = _parse_csv(csv_result["content"])
        csv_headers = list(csv_rows[0].keys()) if csv_rows else list(
            csv.DictReader(io.StringIO(csv_result["content"].decode("utf-8"))).fieldnames
        )
        xlsx_headers, _xlsx_rows = _parse_xlsx(xlsx_result["content"])
        assert xlsx_headers == csv_headers, file_type


@pytest.mark.asyncio
async def test_xlsx_row_count_matches_csv_row_count():
    db = InMemoryDatabase()
    await _seed_finalized_session(db, "sess-19")
    await _seed_default_mappings(db)
    await _seed_erp_item(db, "ITEM-19", is_serialized=True)
    await _seed_approved_line(
        db, "sess-19", "ITEM-19", counted_qty=2.0, erp_qty=2.0, serial_numbers=["SN-A", "SN-B"]
    )
    preview = await ErpNextExportService(db).generate_preview(
        session_id="sess-19", mode="STOCK_ADJUSTMENT", company=DEFAULT_COMPANY, current_user=CURRENT_USER
    )
    await ErpNextExportService(db).approve_preview(
        export_id=preview["export_id"], current_user=ADMIN_USER
    )
    service = ErpNextExportFileService(db)

    csv_result = await service.generate_file(
        export_id=preview["export_id"], file_type="serials", file_format="csv", current_user=ADMIN_USER
    )
    xlsx_result = await service.generate_file(
        export_id=preview["export_id"], file_type="serials", file_format="xlsx", current_user=ADMIN_USER
    )
    csv_rows = _parse_csv(csv_result["content"])
    _headers, xlsx_rows = _parse_xlsx(xlsx_result["content"])

    assert len(csv_rows) == len(xlsx_rows) == 2


@pytest.mark.asyncio
async def test_xlsx_file_hash_matches_generated_bytes():
    import hashlib

    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-20", "ITEM-20")
    service = ErpNextExportFileService(db)

    result = await service.generate_file(
        export_id=approved["export_id"],
        file_type="stock_entry",
        file_format="xlsx",
        current_user=ADMIN_USER,
    )

    assert result["file_hash"] == hashlib.sha256(result["content"]).hexdigest()
    stored = await db.erpnext_export_previews.find_one({"export_id": approved["export_id"]})
    assert stored["file_hashes"]["stock_entry:xlsx"] == result["file_hash"]
    # CSV and XLSX hashes are tracked independently.
    assert "stock_entry:csv" not in stored["file_hashes"]


@pytest.mark.asyncio
async def test_repeated_xlsx_generation_is_idempotent():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-21", "ITEM-21")
    service = ErpNextExportFileService(db)

    first = await service.generate_file(
        export_id=approved["export_id"],
        file_type="stock_entry",
        file_format="xlsx",
        current_user=ADMIN_USER,
    )
    second = await service.generate_file(
        export_id=approved["export_id"],
        file_type="stock_entry",
        file_format="xlsx",
        current_user=ADMIN_USER,
    )

    assert first["file_hash"] == second["file_hash"]
    audit_entries = await db.audit_logs.find(
        {"entity_id": approved["export_id"], "details.file_type": "stock_entry", "details.format": "xlsx"}
    ).to_list(length=10)
    assert len(audit_entries) == 1


@pytest.mark.asyncio
async def test_xlsx_download_content_type_is_correct():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-22", "ITEM-22")
    service = ErpNextExportFileService(db)

    result = await service.generate_file(
        export_id=approved["export_id"],
        file_type="stock_entry",
        file_format="xlsx",
        current_user=ADMIN_USER,
    )

    assert (
        result["media_type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert result["filename"] == "stock-entry.xlsx"


@pytest.mark.asyncio
async def test_csv_endpoints_still_work_after_xlsx_support_added():
    db = InMemoryDatabase()
    approved = await _approve_ready_preview(db, "sess-23", "ITEM-23")
    service = ErpNextExportFileService(db)

    result = await service.generate_file(
        export_id=approved["export_id"], file_type="stock_entry", current_user=ADMIN_USER
    )  # file_format defaults to "csv"

    assert result["media_type"] == "text/csv"
    assert result["filename"] == "stock-entry.csv"
    rows = _parse_csv(result["content"])
    assert rows[0]["Item Code"] == "ITEM-23"


@pytest.mark.asyncio
async def test_audit_event_written_for_first_xlsx_generation_with_format_detail(monkeypatch):
    db = InMemoryDatabase()
    monkeypatch.setattr("backend.api.erpnext_exports_api.get_db", lambda: db)
    approved = await _approve_ready_preview(db, "sess-24", "ITEM-24")

    response = await download_erpnext_export_file(
        approved["export_id"], "stock-entry.xlsx", current_user=CURRENT_USER
    )
    assert (
        response.media_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    entry = await db.audit_logs.find_one(
        {"entity_id": approved["export_id"], "details.format": "xlsx"}
    )
    assert entry is not None
    assert entry["event_type"] == "EXPORT_FILE_GENERATED"
    assert entry["details"]["file_type"] == "stock_entry"


def test_bsr_remediation_status_doc_exists_and_documents_this_step():
    import pathlib

    doc_path = pathlib.Path(__file__).resolve().parents[3] / "docs" / "BSR_REMEDIATION_STATUS.md"
    assert doc_path.exists()
    content = doc_path.read_text()
    for section in (
        "## 1. Current Overall Status",
        "## 2. Non-Negotiable Rules",
        "## 3. Completed Work Log",
        "## 4. Current Verified Capabilities",
        "## 5. Open Gaps",
        "## 6. Step-by-Step Change History",
        "## 7. ERPNext Export Readiness",
        "## 8. Formula Reference",
        "## 9. Required Next Steps",
        "## 10. Agent Instructions",
    ):
        assert section in content, section
    assert "download" in content.lower()
