"""ERPNext import-template validation (BSR).

Read-only checks that a generated CSV/XLSX file (or the approved preview
data it would be built from) actually satisfies what ERPNext's import
template expects, before a human operator is handed the file for manual
import. Never talks to ERPNext, never requires ERPNext credentials, never
mutates the preview -- purely a validation pass.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import load_workbook

from backend.services.erpnext_export_file_service import (
    SLUG_BY_TYPE,
    VALID_FILE_FORMATS,
    VALID_FILE_TYPES,
    build_rows,
    render_file,
)

# Independent source of truth for what ERPNext's import template expects
# per file type -- deliberately not derived from the builder's own
# fieldnames, so a drift between "what we generate" and "what's required"
# is actually detectable rather than trivially always-true.
REQUIRED_COLUMNS: dict[str, list[str]] = {
    "stock_entry": [
        "Company",
        "Stock Entry Type",
        "Is Opening",
        "Item Code",
        "Target Warehouse",
        "Qty",
        "UOM",
        "UOM Conversion Factor",
        "Basic Rate",
        "Batch No",
        "Serial No",
        "Remarks",
    ],
    "stock_reconciliation": [
        "Company",
        "Item Code",
        "Warehouse",
        "Qty",
        "Current Qty",
        "Difference Qty",
        "UOM",
        "UOM Conversion Factor",
        "Valuation Rate",
        "Batch No",
        "Serial No",
        "Remarks",
    ],
    "serials": [
        "Serial No",
        "Item Code",
        "Warehouse",
        "Batch No",
        "Company",
        "Purchase Rate",
        "Status",
    ],
    "batches": ["Batch ID", "Item", "Manufacturing Date", "Expiry Date", "Company"],
    "photo_manifest": [
        "Photo ID",
        "Item Code",
        "Row Key",
        "Source",
        "Storage Kind",
        "Public Reference",
        "Content Hash",
        "Durability Status",
        "Captured At",
        "Captured By",
        "Photo Type",
        "Manifest Hash",
    ],
}

# Row-level fields that must be non-blank for the file type in question.
_REQUIRED_ROW_FIELDS: dict[str, list[str]] = {
    "stock_entry": ["Item Code", "Target Warehouse", "UOM"],
    "stock_reconciliation": ["Item Code", "Warehouse", "UOM"],
    "serials": ["Serial No", "Item Code", "Warehouse"],
    "batches": ["Batch ID", "Item"],
    "photo_manifest": ["Photo ID", "Item Code"],
}


class ErpNextImportValidationError(ValueError):
    """Request-level error (bad file_type/format)."""


def _read_headers(file_format: str, content: bytes) -> list[str]:
    if file_format == "csv":
        reader = csv.reader(io.StringIO(content.decode("utf-8")))
        return next(reader, [])
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(cell) if cell is not None else "" for cell in header_row]


class ErpNextImportValidationService:
    def __init__(self, db: Any) -> None:
        self.db = db

    async def validate(
        self,
        *,
        export_id: str,
        file_type: str,
        file_format: str = "csv",
    ) -> dict[str, Any]:
        if file_type not in VALID_FILE_TYPES:
            raise ErpNextImportValidationError(f"Unknown export file type {file_type!r}")
        if file_format not in VALID_FILE_FORMATS:
            raise ErpNextImportValidationError(f"Unknown export file format {file_format!r}")

        errors: list[str] = []
        warnings: list[str] = []

        preview = await self.db.erpnext_export_previews.find_one({"export_id": export_id})
        if preview is None:
            errors.append("PREVIEW_NOT_FOUND")
            return self._result(
                file_type, file_format, errors, warnings, row_count=0, column_count=0
            )

        if preview.get("status") != "APPROVED":
            errors.append(f"PREVIEW_NOT_APPROVED:{preview.get('status')}")

        company = preview.get("company")
        if not company:
            errors.append("COMPANY_MISSING")

        approval_hash = preview.get("approval_hash")
        if not approval_hash:
            errors.append("APPROVAL_HASH_MISSING")
        else:
            from backend.services.erpnext_export_service import ErpNextExportService

            if ErpNextExportService.compute_approval_hash(preview) != approval_hash:
                errors.append("APPROVAL_HASH_MISMATCH")

        if preview.get("blockers"):
            errors.append("PREVIEW_HAS_BLOCKERS")

        # Row-level safety net: an APPROVED preview must have zero blockers
        # on every row too. These checks are redundant with the guarantees
        # approve_preview() already enforces -- verified directly here
        # rather than merely trusted, since this validation exists
        # specifically to catch any future regression in that guarantee.
        photo_manifest = None
        if file_type == "photo_manifest":
            from backend.services.erpnext_export_photo_manifest_service import (
                ErpNextExportPhotoManifestService,
            )

            photo_manifest = await ErpNextExportPhotoManifestService(self.db).get_manifest(
                export_id
            )

        for row in preview.get("rows") or []:
            item_code = row.get("item_code") or "UNKNOWN"
            row_blockers = row.get("blockers") or []
            if row_blockers:
                errors.append(f"BLOCKED_ROW_EXPORTED:{item_code}:{','.join(row_blockers)}")
            if "UNKNOWN_ITEM_UNMAPPED" in row_blockers:
                errors.append(f"UNKNOWN_ITEM_EXPORTED:{item_code}")
            if "SERIAL_REQUIRED_MISSING" in row_blockers:
                errors.append(f"SERIALIZED_ITEM_WITHOUT_SERIAL:{item_code}")
            if "BATCH_REQUIRED_MISSING" in row_blockers:
                errors.append(f"BATCHED_ITEM_WITHOUT_BATCH:{item_code}")
            if "HIGH_VARIANCE_PHOTO_MISSING" in row_blockers:
                errors.append(f"HIGH_RISK_ITEM_WITHOUT_EVIDENCE:{item_code}")
            if not row.get("erpnext_warehouse"):
                errors.append(f"WAREHOUSE_NOT_MAPPED:{item_code}")
            if not row.get("erpnext_uom"):
                errors.append(f"UOM_NOT_MAPPED:{item_code}")

        # Build the exact same rows the file generator would produce, so
        # column/row checks reflect reality, not an assumption.
        try:
            fieldnames, rows = build_rows(file_type, preview, photo_manifest)
        except Exception:
            fieldnames, rows = [], []

        expected_columns = REQUIRED_COLUMNS[file_type]
        missing_columns = [c for c in expected_columns if c not in fieldnames]
        if missing_columns:
            errors.append(f"MISSING_REQUIRED_COLUMNS:{','.join(missing_columns)}")
        elif fieldnames != expected_columns:
            errors.append("COLUMN_ORDER_MISMATCH")

        required_fields = _REQUIRED_ROW_FIELDS.get(file_type, [])
        for row in rows:
            missing = [f for f in required_fields if not row.get(f)]
            if missing:
                errors.append(f"REQUIRED_VALUE_MISSING:{','.join(missing)}")

        # Render the actual bytes once: verify the real serialized header
        # row matches what was checked above (closing the loop on the
        # serialization layer itself, not just the in-memory fieldnames
        # list), and reuse the same bytes for the file_hash check below.
        content = render_file(file_type, file_format, preview, photo_manifest)
        rendered_headers = _read_headers(file_format, content)
        if fieldnames and rendered_headers != fieldnames:
            errors.append("RENDERED_HEADER_MISMATCH")

        # file_hash check: only meaningful if the file has already been
        # generated at least once; otherwise this is informational, not a
        # blocking condition (a file can be validated before it's ever
        # been downloaded).
        hash_key = f"{file_type}:{file_format}"
        stored_hash = (preview.get("file_hashes") or {}).get(hash_key)
        if stored_hash:
            import hashlib

            recomputed = hashlib.sha256(content).hexdigest()
            if recomputed != stored_hash:
                errors.append("FILE_HASH_MISMATCH")
        else:
            warnings.append("FILE_NOT_YET_GENERATED")

        return self._result(
            file_type,
            file_format,
            errors,
            warnings,
            row_count=len(rows),
            column_count=len(fieldnames),
        )

    @staticmethod
    def _result(
        file_type: str,
        file_format: str,
        errors: list[str],
        warnings: list[str],
        *,
        row_count: int,
        column_count: int,
    ) -> dict[str, Any]:
        return {
            "valid": not errors,
            "file_type": SLUG_BY_TYPE.get(file_type, file_type),
            "format": file_format,
            "errors": errors,
            "warnings": warnings,
            "row_count": row_count,
            "column_count": column_count,
        }
