"""ERPNext official import-template validation (BSR).

Compares Stock Verify's generated CSV/XLSX headers against a template
*actually downloaded from the target ERPNext instance* (Setup > Data Import
> Download Template), not against Stock Verify's own internal
`REQUIRED_COLUMNS` spec (see erpnext_export_import_validation_service.py).
Internal-spec validation proves self-consistency; this proves (or disproves)
that ERPNext itself would recognize the columns.

If no template/manifest has been placed in the configured directory, this
never fakes a result -- it reports `ERP_TEMPLATE_SOURCE_MISSING` /
`ERP_TEMPLATE_MANIFEST_MISSING` and falls back to labeling the result
`STOCK_VERIFY_INTERNAL_SPEC` so callers can tell the two apart. Never writes
to ERPNext, never requires credentials, read-only against the local
filesystem.
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

# Anchored to the repo root so the default resolves regardless of the
# process working directory (pytest may run from backend/ or the repo root).
DEFAULT_TEMPLATE_DIR = Path(__file__).resolve().parents[2] / "docs" / "erpnext_templates"

# Maps Stock Verify's internal file_type to the manifest's "templates" key.
# photo_manifest has no entry: it is not an ERPNext import target doctype
# (an audit artifact only), so no ERPNext template is ever expected for it.
_TEMPLATE_KEY_BY_FILE_TYPE = {
    "stock_entry": "stock_entry",
    "stock_reconciliation": "stock_reconciliation",
    "serials": "serial_no",
    "batches": "batch",
}


def _is_unset(value: Optional[str]) -> bool:
    """True for a genuinely missing value AND for the explicit "unknown"
    sentinel a manifest scaffold uses to honestly document "not yet
    confirmed" (see docs/erpnext_templates/template_manifest.json) --
    "unknown" is a truthy non-empty string, so a plain falsy-check would
    miss it."""
    return not value or str(value).strip().lower() == "unknown"


def _read_csv_headers(path: Path) -> list[str]:
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        return next(reader, [])


def _read_xlsx_headers(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(cell) if cell is not None else "" for cell in row]


class ErpNextTemplateValidationService:
    def __init__(self, template_dir: Optional[Path] = None, *, strict: bool = False) -> None:
        self.template_dir = Path(template_dir) if template_dir else DEFAULT_TEMPLATE_DIR
        self.strict = strict

    def load_manifest(self) -> Optional[dict[str, Any]]:
        manifest_path = self.template_dir / "template_manifest.json"
        if not manifest_path.exists():
            return None
        return json.loads(manifest_path.read_text(encoding="utf-8"))

    def get_target_instance_metadata(self) -> dict[str, Any]:
        """Part K/Step 11: target-instance metadata used to gate serial/batch
        opening-stock behavior. `uses_serial_batch_bundle`,
        `negative_serial_batch_stock_allowed`, `serial_import_path`, and
        `batch_import_path` live under the manifest's `serial_batch_gate`
        object (Step 11 schema); a bare top-level
        `uses_serial_batch_bundle`/`negative_serial_batch_stock_allowed`
        (the earlier Step 10 shape) is still honored as a fallback so an
        older manifest isn't silently ignored. Read as literal
        `True`/`False`/`"unknown"` (or absent) -- never inferred."""
        manifest = self.load_manifest() or {}
        gate = manifest.get("serial_batch_gate") or {}
        return {
            "erpnext_version": manifest.get("erpnext_version") or None,
            "frappe_version": manifest.get("frappe_version") or None,
            "uses_serial_batch_bundle": gate.get(
                "uses_serial_batch_bundle", manifest.get("uses_serial_batch_bundle", "unknown")
            ),
            "negative_serial_batch_stock_allowed": gate.get(
                "negative_serial_batch_stock_allowed",
                manifest.get("negative_serial_batch_stock_allowed", "unknown"),
            ),
            "serial_import_path": gate.get("serial_import_path", "unknown"),
            "batch_import_path": gate.get("batch_import_path", "unknown"),
        }

    @staticmethod
    def _major_version(version_string: Optional[str]) -> Optional[int]:
        if _is_unset(version_string):
            return None
        match = re.search(r"v?(\d+)", str(version_string))
        return int(match.group(1)) if match else None

    def evaluate_serial_batch_gate(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        """Part K: ERPNext Serial and Batch Bundle gate. Never hardcodes v15
        behavior as universal -- the stricter gate only applies when the
        manifest actually says the target is v15+ or
        `uses_serial_batch_bundle` is true. When target-instance metadata is
        genuinely unknown, this blocks rather than assumes either way (per
        the explicit rule: "If target version is unknown, final manual-
        import readiness remains blocked."). Uses `serial_numbers`/
        `batch_no` presence on the frozen row as the serialized/batch-
        controlled signal -- the same fields already frozen at preview time,
        no new data source.
        """
        metadata = self.get_target_instance_metadata()
        major = self._major_version(metadata["erpnext_version"])
        uses_bundle = metadata["uses_serial_batch_bundle"]
        version_known = major is not None or uses_bundle in (True, False, "true", "false")
        gate_applies = (major is not None and major >= 15) or uses_bundle in (True, "true")

        blockers: list[dict[str, str]] = []
        warnings: list[str] = []

        if not version_known:
            blockers.append(
                {
                    "code": "ERP_VERSION_UNKNOWN",
                    "message": "Target ERPNext version / Serial and Batch Bundle usage is "
                    "unknown; manual-import readiness cannot proceed past "
                    "READY_FOR_ERP_TEMPLATE_COMPARISON until this is confirmed.",
                }
            )

        for row in rows:
            is_serialized_row = bool(row.get("serial_numbers"))
            is_batch_row = bool(row.get("batch_no"))
            if not (is_serialized_row or is_batch_row):
                continue
            qty = row.get("erpnext_opening_qty")
            if qty is None or qty >= 0:
                continue
            item_code = row.get("item_code")
            if not version_known:
                blockers.append(
                    {
                        "code": "SERIAL_BATCH_NEGATIVE_QTY_VERSION_UNKNOWN",
                        "message": f"Item {item_code!r} has a negative serialized/batch "
                        "opening quantity and the target ERPNext version is unknown.",
                    }
                )
            elif gate_applies:
                blockers.append(
                    {
                        "code": "SERIAL_BATCH_NEGATIVE_QTY_NEEDS_REVIEW",
                        "message": f"Item {item_code!r} has a negative serialized/batch "
                        "opening quantity on a target using Serial and Batch Bundle; "
                        "requires explicit review/override before import.",
                    }
                )
            else:
                warnings.append(f"NEGATIVE_SERIAL_BATCH_QTY:{item_code}")

        return {
            "metadata": metadata,
            "gate_applies": gate_applies,
            "version_known": version_known,
            "blockers": blockers,
            "warnings": warnings,
        }

    def validate(
        self,
        *,
        file_type: str,
        file_format: str,
        stock_verify_headers: list[str],
        strict: Optional[bool] = None,
    ) -> dict[str, Any]:
        strict = self.strict if strict is None else strict
        errors: list[dict[str, str]] = []
        warnings: list[str] = []
        erpnext_headers: list[str] = []
        template_available = False

        manifest = self.load_manifest()
        if manifest is None:
            errors.append(
                {
                    "code": "ERP_TEMPLATE_MANIFEST_MISSING",
                    "message": "No ERPNext template manifest found; templates have not been "
                    "downloaded from the target instance (see docs/erpnext_templates/README.md).",
                }
            )
        else:
            if _is_unset(manifest.get("erpnext_version")):
                errors.append(
                    {
                        "code": "ERP_VERSION_UNKNOWN",
                        "message": "Template manifest does not record which ERPNext version "
                        "these templates came from.",
                    }
                )

            template_key = _TEMPLATE_KEY_BY_FILE_TYPE.get(file_type)
            template_entry = (manifest.get("templates") or {}).get(template_key) if template_key else None

            if file_type == "photo_manifest":
                pass  # Not an ERPNext import target doctype; no template expected.
            elif template_entry is None:
                errors.append(
                    {
                        "code": "ERP_TEMPLATE_SOURCE_MISSING",
                        "message": f"No target ERPNext template file is available for {file_type!r}.",
                    }
                )
            else:
                template_file = self.template_dir / template_entry["file"]
                if not template_file.exists():
                    errors.append(
                        {
                            "code": "ERP_TEMPLATE_SOURCE_MISSING",
                            "message": f"No target ERPNext template file is available for {file_type!r}.",
                        }
                    )
                else:
                    erpnext_headers = (
                        _read_csv_headers(template_file)
                        if template_file.suffix == ".csv"
                        else _read_xlsx_headers(template_file)
                    )
                    template_available = True
                    if template_entry.get("has_child_table") and template_entry.get(
                        "child_table_name"
                    ) in (None, "unknown"):
                        warnings.append("CHILD_TABLE_AMBIGUITY")

        missing_columns: list[str] = []
        extra_columns: list[str] = []
        case_mismatches: list[dict[str, str]] = []
        order_mismatches: list[dict[str, list[str]]] = []

        if template_available:
            sv_lower = {h.lower(): h for h in stock_verify_headers}
            erp_lower = {h.lower(): h for h in erpnext_headers}

            for header in erpnext_headers:
                if header in stock_verify_headers:
                    continue
                if header.lower() in sv_lower:
                    case_mismatches.append({"erpnext": header, "stock_verify": sv_lower[header.lower()]})
                else:
                    missing_columns.append(header)

            for header in stock_verify_headers:
                if header in erpnext_headers or header.lower() in erp_lower:
                    continue
                extra_columns.append(header)

            common_in_erpnext_order = [h for h in erpnext_headers if h in stock_verify_headers]
            common_in_stock_verify_order = [h for h in stock_verify_headers if h in erpnext_headers]
            if common_in_erpnext_order != common_in_stock_verify_order:
                order_mismatches.append(
                    {"expected_order": common_in_erpnext_order, "actual_order": common_in_stock_verify_order}
                )

            if missing_columns:
                errors.append(
                    {
                        "code": "ERP_TEMPLATE_MISSING_COLUMNS",
                        "message": f"Columns required by the ERPNext template are missing: {missing_columns}",
                    }
                )
            if case_mismatches:
                errors.append(
                    {
                        "code": "ERP_TEMPLATE_CASE_MISMATCH",
                        "message": f"Header case differs from the ERPNext template: {case_mismatches}",
                    }
                )
            if order_mismatches:
                errors.append(
                    {
                        "code": "ERP_TEMPLATE_ORDER_MISMATCH",
                        "message": "Column order differs from the ERPNext template.",
                    }
                )
            if extra_columns:
                if strict:
                    errors.append(
                        {
                            "code": "ERP_TEMPLATE_EXTRA_COLUMNS_STRICT",
                            "message": f"Columns not present in the ERPNext template: {extra_columns}",
                        }
                    )
                else:
                    warnings.append(f"EXTRA_COLUMNS_NOT_IN_TEMPLATE:{','.join(extra_columns)}")

        return {
            "valid": not errors,
            "source": "ERP_NEXT_TEMPLATE" if template_available else "STOCK_VERIFY_INTERNAL_SPEC",
            "template_available": template_available,
            "file_type": file_type,
            "format": file_format,
            "errors": errors,
            "warnings": warnings,
            "erpnext_headers": erpnext_headers,
            "stock_verify_headers": list(stock_verify_headers),
            "missing_columns": missing_columns,
            "extra_columns": extra_columns,
            "case_mismatches": case_mismatches,
            "order_mismatches": order_mismatches,
        }
