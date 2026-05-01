"""
Report Generation API - Multiple report types with filtering and export
Supports Stock Summary, Variance, User Activity, Session History, and Audit Trail reports
"""

import csv
import io
import json
import logging
from datetime import date, datetime, timezone
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from pymongo.errors import PyMongoError

from backend.auth.dependencies import get_current_user, require_role
from backend.services.query_utils import build_mongo_date_filter
from backend.services.report_generation_service import (
    ReportGenerationService,
    get_report_generation_service,
)
from backend.utils.api_utils import sanitize_for_logging
from backend.utils.request_context import get_request_id

logger = logging.getLogger(__name__)

report_generation_router = APIRouter(prefix="/reports", tags=["Reports"])


# Report Types
REPORT_TYPES = {
    "stock_summary": {
        "name": "Stock Summary Report",
        "description": "Overview of stock levels, values, and verification status",
        "collection": "count_lines",
    },
    "variance_report": {
        "name": "Variance Report",
        "description": "Items with discrepancies between expected and counted quantities",
        "collection": "count_lines",
    },
    "user_activity": {
        "name": "User Activity Report",
        "description": "User actions, sessions, and productivity metrics",
        "collection": "audit_logs",
    },
    "session_history": {
        "name": "Session History Report",
        "description": "Verification session details and outcomes",
        "collection": "sessions",
    },
    "audit_trail": {
        "name": "Audit Trail Report",
        "description": "Complete audit log of all system actions",
        "collection": "audit_logs",
    },
}


# Request/Response Models
class ReportFilter(BaseModel):
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    warehouse: Optional[str] = None
    user_id: Optional[str] = None
    status: Optional[str] = None
    floor: Optional[str] = None
    category: Optional[str] = None


class ReportRequest(BaseModel):
    report_type: str
    filters: Optional[ReportFilter] = None
    format: str = Field(default="json", pattern="^(json|csv|xlsx)$")
    include_summary: bool = True


class ReportSummary(BaseModel):
    total_records: int
    generated_at: str
    filters_applied: dict[str, Any]
    report_type: str
    report_name: str


class ReportResponse(BaseModel):
    summary: ReportSummary
    data: list[dict[str, Any]]


# Helper Functions
def build_date_filter(
    date_from: Optional[date], date_to: Optional[date]
) -> Optional[dict[str, Any]]:
    """Build MongoDB date range filter."""
    return build_mongo_date_filter(date_from, date_to, end_of_day=True)


def sanitize_for_csv(value: Any) -> str:
    """Sanitize value for CSV export."""
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _format_xlsx_cell_value(value: Any) -> Any:
    """Format a value for Excel cell."""
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _write_xlsx_headers(ws: Any, headers: list[str]) -> None:
    """Write header row to Excel worksheet."""
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)


def _write_xlsx_data(ws: Any, data: list[dict], headers: list[str]) -> None:
    """Write data rows to Excel worksheet."""
    for row_idx, row in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row.get(header)
            ws.cell(row=row_idx, column=col_idx, value=_format_xlsx_cell_value(value))


async def generate_stock_summary(db, filters: ReportFilter) -> list[dict]:
    """Generate stock summary report data."""
    return await ReportGenerationService(db).generate_stock_summary(filters)


async def generate_variance_report(db, filters: ReportFilter) -> list[dict]:
    """Generate variance report data."""
    return await ReportGenerationService(db).generate_variance_report(filters)


async def generate_user_activity_report(db, filters: ReportFilter) -> list[dict]:
    """Generate user activity report data."""
    return await ReportGenerationService(db).generate_user_activity_report(filters)


async def generate_session_history_report(db, filters: ReportFilter) -> list[dict]:
    """Generate session history report data."""
    return await ReportGenerationService(db).generate_session_history_report(filters)


async def generate_audit_trail_report(db, filters: ReportFilter) -> list[dict]:
    """Generate audit trail report data."""
    return await ReportGenerationService(db).generate_audit_trail_report(filters)


# Report Generator Dispatch
REPORT_GENERATORS = {
    "stock_summary": generate_stock_summary,
    "variance_report": generate_variance_report,
    "user_activity": generate_user_activity_report,
    "session_history": generate_session_history_report,
    "audit_trail": generate_audit_trail_report,
}


# API Endpoints
@report_generation_router.get("/types")
async def get_report_types(current_user: dict = Depends(get_current_user)):
    """Get available report types with descriptions."""
    return {
        "report_types": [
            {"id": key, "name": value["name"], "description": value["description"]}
            for key, value in REPORT_TYPES.items()
        ]
    }


@report_generation_router.post("/generate", response_model=ReportResponse)
async def generate_report(
    report_request: ReportRequest,
    http_request: Request,
    current_user: dict = Depends(require_role("admin", "supervisor")),
    report_service: ReportGenerationService = Depends(get_report_generation_service),
):
    """
    Generate a report with specified filters.
    Returns data in JSON format by default.
    """
    request_id = get_request_id(http_request.headers)
    if report_request.report_type not in REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid report type. Valid types: {list(REPORT_TYPES.keys())}",
        )

    filters = report_request.filters or ReportFilter()

    # Generate report data
    generator = REPORT_GENERATORS[report_request.report_type]
    try:
        data = await getattr(report_service, generator.__name__)(filters)
    except (KeyError, PyMongoError, RuntimeError, TypeError, ValueError) as e:
        logger.error(
            "report_generation_failed",
            extra={
                "request_id": request_id,
                "user": sanitize_for_logging(
                    str(current_user.get("username", "unknown"))
                ),
                "endpoint": "/reports/generate",
                "status": "failed",
                "report_type": sanitize_for_logging(report_request.report_type),
                "error": sanitize_for_logging(str(e)),
            },
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate report",
        )

    # Build summary
    filters_applied = {k: v for k, v in filters.model_dump().items() if v is not None}

    summary = ReportSummary(
        total_records=len(data),
        generated_at=datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
        filters_applied=filters_applied,
        report_type=report_request.report_type,
        report_name=REPORT_TYPES[report_request.report_type]["name"],
    )
    logger.info(
        "report_generated",
        extra={
            "request_id": request_id,
            "user": sanitize_for_logging(str(current_user.get("username", "unknown"))),
            "endpoint": "/reports/generate",
            "status": "completed",
            "report_type": sanitize_for_logging(report_request.report_type),
            "total_records": len(data),
        },
    )

    return ReportResponse(summary=summary, data=data)


@report_generation_router.post("/export/csv")
async def export_report_csv(
    request: ReportRequest,
    current_user: dict = Depends(require_role("admin", "supervisor")),
    report_service: ReportGenerationService = Depends(get_report_generation_service),
):
    """
    Export report as CSV file.
    """
    if request.report_type not in REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid report type"
        )

    filters = request.filters or ReportFilter()

    # Generate report data
    generator = REPORT_GENERATORS[request.report_type]
    try:
        data = await getattr(report_service, generator.__name__)(filters)
    except (KeyError, PyMongoError, RuntimeError, TypeError, ValueError) as e:
        logger.error("Error generating report: %s", sanitize_for_logging(str(e)))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate report",
        )

    if not data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No data found for the specified filters",
        )

    # Create CSV
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()

    for row in data:
        sanitized_row = {k: sanitize_for_csv(v) for k, v in row.items()}
        writer.writerow(sanitized_row)

    output.seek(0)

    filename = f"{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@report_generation_router.post("/export/xlsx")
async def export_report_xlsx(
    request: ReportRequest,
    current_user: dict = Depends(require_role("admin", "supervisor")),
    report_service: ReportGenerationService = Depends(get_report_generation_service),
):
    """
    Export report as Excel XLSX file.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export not available. Install openpyxl package.",
        )

    if request.report_type not in REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid report type"
        )

    filters = request.filters or ReportFilter()

    generator = REPORT_GENERATORS[request.report_type]
    try:
        data = await getattr(report_service, generator.__name__)(filters)
    except (KeyError, PyMongoError, RuntimeError, TypeError, ValueError) as e:
        logger.error("Error generating report: %s", sanitize_for_logging(str(e)))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate report",
        )

    if not data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No data found for the specified filters",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = REPORT_TYPES[request.report_type]["name"][:31]

    headers = list(data[0].keys())
    _write_xlsx_headers(ws, headers)
    _write_xlsx_data(ws, data, headers)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@report_generation_router.get("/filters/{report_type}")
async def get_report_filter_options(
    report_type: str,
    current_user: dict = Depends(get_current_user),
    report_service: ReportGenerationService = Depends(get_report_generation_service),
):
    """
    Get available filter options for a specific report type.
    Returns distinct values for filterable fields.
    """
    if report_type not in REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid report type"
        )

    try:
        return {
            "report_type": report_type,
            "filters": await report_service.get_filter_options(current_user),
        }

    except (KeyError, PyMongoError, RuntimeError, TypeError, ValueError) as e:
        logger.error("Error fetching filter options: %s", sanitize_for_logging(str(e)))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch filter options",
        )
