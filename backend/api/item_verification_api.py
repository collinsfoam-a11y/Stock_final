"""
Item Verification API - Verification, filtering, CSV export, and variance tracking
"""

import asyncio
import csv
import io
import json
import logging
from copy import deepcopy
from datetime import datetime, timezone
from typing import Any, Optional, cast

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorDatabase
from pymongo.errors import PyMongoError
from pydantic import BaseModel

from backend.auth.dependencies import get_current_user_async as get_current_user
from backend.utils.api_utils import sanitize_for_logging
from backend.services.governance_guard import write_authority
from backend.services.item_verification_service import ItemVerificationService

logger = logging.getLogger(__name__)

# These will be initialized at runtime
db: AsyncIOMotorDatabase = cast(AsyncIOMotorDatabase, None)
cache_service: Any = None
sql_sync_service: Any = None
ITEM_VERIFICATION_ERRORS = (KeyError, PyMongoError, RuntimeError, TypeError, ValueError)


def init_verification_api(
    database: AsyncIOMotorDatabase, cache_svc: Any = None, sql_svc: Any = None
) -> None:
    """Initialize verification API with dependencies"""
    global db, cache_service, sql_sync_service
    db = database
    cache_service = cache_svc
    sql_sync_service = sql_svc


verification_router = APIRouter(prefix="/api/v2/erp/items", tags=["Item Verification"])


def _item_service() -> ItemVerificationService:
    return ItemVerificationService(db)


def _safe_log_value(value: Any, *, max_length: int = 120) -> str:
    """Return a log-safe string for user-controlled values and exception messages."""
    return sanitize_for_logging("" if value is None else str(value), max_length=max_length)


def _regex_filter(value: Optional[str]) -> Optional[dict[str, str]]:
    if not value:
        return None
    return {"$regex": value, "$options": "i"}


def build_item_filter_query(
    *,
    category: Optional[str] = None,
    subcategory: Optional[str] = None,
    floor: Optional[str] = None,
    rack: Optional[str] = None,
    warehouse: Optional[str] = None,
    uom_code: Optional[str] = None,
    verified: Optional[bool] = None,
    search: Optional[str] = None,
) -> dict[str, Any]:
    """Create a MongoDB filter dict for ERP items."""
    filter_query: dict[str, Any] = {}

    category_filter = _regex_filter(category)
    if category_filter:
        filter_query["category"] = category_filter

    subcategory_filter = _regex_filter(subcategory)
    if subcategory_filter:
        filter_query["subcategory"] = subcategory_filter

    floor_filter = _regex_filter(floor)
    if floor_filter:
        filter_query["floor"] = floor_filter

    rack_filter = _regex_filter(rack)
    if rack_filter:
        filter_query["rack"] = rack_filter

    warehouse_filter = _regex_filter(warehouse)
    if warehouse_filter:
        filter_query["warehouse"] = warehouse_filter

    if uom_code:
        filter_query["uom_code"] = uom_code

    if verified is not None:
        filter_query["verified"] = verified

    if search:
        filter_query["$or"] = [
            {"item_name": {"$regex": search, "$options": "i"}},
            {"item_code": {"$regex": search, "$options": "i"}},
            {"barcode": {"$regex": search, "$options": "i"}},
        ]

    return filter_query


def serialize_mongo_datetime(value: Optional[datetime]) -> str:
    return value.isoformat() if isinstance(value, datetime) else ""


def _serialize_export_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _render_xlsx_bytes(fieldnames: list[str], rows: list[dict[str, Any]]) -> bytes:
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Install openpyxl package.",
        ) from exc

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if sheet is None:
        sheet = workbook.create_sheet("Export")

    for column_index, field in enumerate(fieldnames, 1):
        sheet.cell(row=1, column=column_index, value=field)

    for row_index, row in enumerate(rows, 2):
        for column_index, field in enumerate(fieldnames, 1):
            sheet.cell(
                row=row_index,
                column=column_index,
                value=_serialize_export_value(row.get(field, "")),
            )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


ITEM_EXPORT_FIELDNAMES = [
    "ID",
    "item_code",
    "item_name",
    "barcode",
    "stock_qty",
    "mrp",
    "category",
    "subcategory",
    "uom_code",
    "uom_name",
    "floor",
    "rack",
    "warehouse",
    "verified",
    "verified_by",
    "verified_at",
    "last_scanned_at",
    "verified_qty",
    "damaged_qty",
    "non_returnable_damaged_qty",
    "variance",
    "item_condition",
    "serial_number",
    "is_serialized",
    "session_id",
    "verification_notes",
]


VARIANCE_EXPORT_FIELDNAMES = [
    "ID",
    "item_code",
    "item_name",
    "system_qty",
    "verified_qty",
    "variance",
    "verified_by",
    "verified_at",
    "category",
    "subcategory",
    "floor",
    "rack",
    "warehouse",
    "session_id",
    "count_line_id",
]


def _build_erpnext_item_export_row(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "ID": "",
        "item_code": item.get("item_code", ""),
        "item_name": item.get("item_name", ""),
        "barcode": item.get("barcode", ""),
        "stock_qty": item.get("stock_qty", 0.0),
        "mrp": item.get("mrp", 0.0),
        "category": item.get("category", ""),
        "subcategory": item.get("subcategory", ""),
        "uom_code": item.get("uom_code", ""),
        "uom_name": item.get("uom_name", ""),
        "floor": item.get("floor", ""),
        "rack": item.get("rack", ""),
        "warehouse": item.get("warehouse", ""),
        "verified": "Yes" if item.get("verified", False) else "No",
        "verified_by": item.get("verified_by", ""),
        "verified_at": serialize_mongo_datetime(item.get("verified_at")),
        "last_scanned_at": serialize_mongo_datetime(item.get("last_scanned_at")),
        "verified_qty": item.get("verified_qty", 0.0),
        "damaged_qty": item.get("damaged_qty", 0.0),
        "non_returnable_damaged_qty": item.get("non_returnable_damaged_qty", 0.0),
        "variance": item.get("variance", 0.0),
        "item_condition": item.get("item_condition", ""),
        "serial_number": item.get("serial_number", ""),
        "is_serialized": "Yes" if item.get("is_serialized", False) else "No",
        "session_id": item.get("session_id", ""),
        "verification_notes": item.get("verification_notes", ""),
    }


def _build_erpnext_variance_export_row(variance: dict[str, Any]) -> dict[str, Any]:
    return {
        "ID": "",
        "item_code": variance.get("item_code", ""),
        "item_name": variance.get("item_name", ""),
        "system_qty": variance.get("system_qty", 0.0),
        "verified_qty": variance.get("verified_qty", 0.0),
        "variance": variance.get("variance", 0.0),
        "verified_by": variance.get("verified_by", ""),
        "verified_at": serialize_mongo_datetime(variance.get("verified_at")),
        "category": variance.get("category", ""),
        "subcategory": variance.get("subcategory", ""),
        "floor": variance.get("floor", ""),
        "rack": variance.get("rack", ""),
        "warehouse": variance.get("warehouse", ""),
        "session_id": variance.get("session_id", ""),
        "count_line_id": variance.get("count_line_id", ""),
    }


def serialize_item_document(item: dict[str, Any]) -> dict[str, Any]:
    """Return a JSON-serializable copy of an ERP item document."""
    serialized = dict(item)
    if "_id" in serialized:
        serialized["_id"] = str(serialized["_id"])

    for field in ("verified_at", "last_scanned_at"):
        if field in serialized:
            serialized[field] = serialize_mongo_datetime(serialized.get(field))

    return serialized


class VerificationRequest(BaseModel):
    verified: bool
    verified_qty: Optional[float] = None
    damaged_qty: Optional[float] = 0.0
    non_returnable_damaged_qty: Optional[float] = 0.0
    item_condition: Optional[str] = "Good"
    serial_number: Optional[str] = None
    # Deprecated for writes: serialization policy is master-data controlled.
    is_serialized: Optional[bool] = None
    notes: Optional[str] = None
    floor: Optional[str] = None
    rack: Optional[str] = None
    session_id: Optional[str] = None
    count_line_id: Optional[str] = None


class ItemUpdateRequest(BaseModel):
    mrp: Optional[float] = None
    sales_price: Optional[float] = None
    category: Optional[str] = None
    subcategory: Optional[str] = None
    uom: Optional[str] = None


def _not_found_error(barcode: str) -> HTTPException:
    return HTTPException(status_code=404, detail=f"Item with barcode/code {barcode} not found")


async def _find_item_by_barcode_or_code(barcode: str) -> dict[str, Any]:
    item = await _item_service().find_item_by_barcode_or_code(barcode)
    if item:
        return item

    raise _not_found_error(barcode)


def _resolve_item_identity(
    item: dict[str, Any], default_barcode: str
) -> tuple[Optional[str], str, dict[str, Any]]:
    actual_barcode = item.get("barcode")
    actual_item_code = item.get("item_code") or default_barcode
    if actual_item_code:
        return actual_barcode, actual_item_code, {"item_code": actual_item_code}
    if actual_barcode:
        return actual_barcode, actual_item_code, {"barcode": actual_barcode}
    return actual_barcode, actual_item_code, {"barcode": default_barcode}


def _build_master_update_doc(
    request: ItemUpdateRequest, current_user: dict[str, Any]
) -> dict[str, Any]:
    update_fields: dict[str, Any] = {
        "last_updated_by": current_user["username"],
        "last_updated_at": datetime.now(timezone.utc).replace(tzinfo=None),
    }
    field_mapping = {
        "mrp": "mrp",
        "sales_price": "sales_price",
        "category": "category",
        "subcategory": "subcategory",
        "uom": "uom",
    }
    for req_field, update_field in field_mapping.items():
        value = getattr(request, req_field)
        if value is not None:
            update_fields[update_field] = value
    return {"$set": update_fields}


async def _invalidate_item_cache(
    *,
    actual_barcode: Optional[str],
    actual_item_code: Optional[str],
    clear_search_cache: bool = False,
) -> None:
    if not cache_service:
        return
    if actual_barcode:
        await cache_service.delete_async("items", f"enhanced_{actual_barcode}")
    if actual_item_code:
        await cache_service.delete_async("items", f"enhanced_{actual_item_code}")
    if clear_search_cache:
        await cache_service.clear_pattern("search:*")


async def _insert_master_update_audit_log(
    *,
    actual_item_code: str,
    actual_barcode: Optional[str],
    requested_barcode: str,
    request: ItemUpdateRequest,
    current_user: dict[str, Any],
) -> None:
    await _item_service().insert_master_update_audit_log(
        {
            "action": "MASTER_UPDATE",
            "item_code": actual_item_code,
            "barcode": actual_barcode or requested_barcode,
            "changes": request.model_dump(exclude_none=True),
            "user": current_user["username"],
            "timestamp": datetime.now(timezone.utc).replace(tzinfo=None),
        }
    )


def _build_verification_filter(
    *,
    actual_barcode: Optional[str],
    actual_item_code: str,
    requested_barcode: str,
    expected_stock_qty: Optional[float],
) -> dict[str, Any]:
    if actual_item_code:
        update_filter: dict[str, Any] = {"item_code": actual_item_code}
    elif actual_barcode:
        update_filter = {"barcode": actual_barcode}
    else:
        update_filter = {"barcode": requested_barcode}

    if expected_stock_qty is not None:
        update_filter["stock_qty"] = expected_stock_qty
    return update_filter


async def _fetch_item_with_optional_sql_refresh(barcode: str) -> dict[str, Any]:
    if sql_sync_service:
        try:
            if sql_sync_service.sql_connector.test_connection():
                refreshed_item = await sql_sync_service.sync_single_item_by_barcode(barcode)
                if refreshed_item:
                    return refreshed_item
        except (RuntimeError, OSError, TypeError, ValueError) as e:
            logger.warning(
                "Failed to auto-refresh item %s from SQL: %s",
                sanitize_for_logging(barcode),
                sanitize_for_logging(str(e)),
            )
    return await _find_item_by_barcode_or_code(barcode)


def _verification_has_material_conflict(item: dict[str, Any], request: VerificationRequest) -> bool:
    if item.get("verified") is not True:
        return False
    existing_qty = item.get("verified_qty")
    return request.verified_qty is not None and request.verified_qty != existing_qty


async def _create_conflict_fork_response(
    *,
    item: dict[str, Any],
    request: VerificationRequest,
    current_user: dict[str, Any],
    barcode: str,
) -> dict[str, Any]:
    from backend.core.schemas.conflict import ConflictFork

    existing_qty = item.get("verified_qty")
    fork = ConflictFork(
        original_item_id=str(item.get("_id")),
        session_id=request.session_id or "unknown",
        user_id=current_user["username"],
        conflicting_payload=request.model_dump(exclude_none=True),
        reason=f"Attempted to overwrite APPROVED qty {existing_qty} with {request.verified_qty}",
    )

    await _item_service().insert_conflict_fork(fork.model_dump())
    logger.warning(
        "Conflict detected for %s. Fork created: %s",
        _safe_log_value(barcode),
        _safe_log_value(fork.fork_id),
    )
    return {
        "success": True,
        "item": serialize_item_document(item),
        "variance": item.get("variance"),
        "message": f"Conflict detected! Original verification preserved. Fork ID: {fork.fork_id}",
        "fork_id": fork.fork_id,
    }


async def _fetch_updated_item(
    update_filter: dict[str, Any], actual_barcode: Optional[str]
) -> dict[str, Any]:
    updated_item = await _item_service().fetch_updated_item(update_filter, actual_barcode)
    if not updated_item:
        raise HTTPException(status_code=500, detail="Verification updated item not found")
    updated_item["_id"] = str(updated_item["_id"])
    return updated_item


@verification_router.patch("/{barcode}/update-master")
async def update_item_master(
    barcode: str,
    request: ItemUpdateRequest,
    current_user: dict = Depends(get_current_user),
):
    """
    Update item master details (MRP, Price, Category, etc.)
    """
    try:
        item = await _find_item_by_barcode_or_code(barcode)
        actual_barcode, actual_item_code, update_filter = _resolve_item_identity(item, barcode)
        with write_authority("ItemVerificationAPI"):
            await _item_service().update_erp_item(
                update_filter,
                _build_master_update_doc(request, current_user),
            )
        await _invalidate_item_cache(
            actual_barcode=actual_barcode,
            actual_item_code=actual_item_code,
            clear_search_cache=True,
        )
        await _insert_master_update_audit_log(
            actual_item_code=actual_item_code,
            actual_barcode=actual_barcode,
            requested_barcode=barcode,
            request=request,
            current_user=current_user,
        )
        return {"success": True, "message": "Item details updated successfully"}

    except HTTPException:
        raise
    except ITEM_VERIFICATION_ERRORS as e:
        logger.error(
            "Error updating item master %s: %s",
            _safe_log_value(barcode),
            _safe_log_value(e, max_length=200),
        )
        raise HTTPException(status_code=500, detail=str(e))


@verification_router.post("/{barcode}/refresh-sql-qty")
async def refresh_item_qty_from_sql(
    barcode: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Manually refresh item quantity from SQL Server.
    Updates MongoDB if there's a difference.
    """
    if not sql_sync_service:
        raise HTTPException(
            status_code=503,
            detail="SQL Sync Service not available (SQL Server might be disabled or offline)",
        )

    try:
        # Check if SQL is connected (wrapped to avoid blocking event loop)
        is_connected = await asyncio.to_thread(sql_sync_service.sql_connector.test_connection)
        if not is_connected:
            raise HTTPException(status_code=503, detail="SQL Server is currently unreachable")

        # Attempt to sync single item
        # Try finding by barcode first, as the param suggests
        updated_item = await sql_sync_service.sync_single_item_by_barcode(barcode)

        if not updated_item:
            # If not found by barcode, try as item_code (legacy support)
            # This requires looking up the item code first or trusting the caller
            # For now, we assume barcode is barcode.
            # If we want to support item_code, we might need a separate service method
            # or try to interpret the barcode.
            pass

        if not updated_item:
            raise HTTPException(
                status_code=404, detail=f"Item with barcode {barcode} not found in SQL Server"
            )

        updated_item["_id"] = str(updated_item["_id"])

        # Invalidate cache
        if cache_service:
            if updated_item.get("barcode"):
                await cache_service.delete_async("items", f"enhanced_{updated_item['barcode']}")
            if updated_item.get("item_code"):
                await cache_service.delete_async("items", f"enhanced_{updated_item['item_code']}")

        return {
            "success": True,
            "message": "Quantity refreshed from SQL Server",
            "item": serialize_item_document(updated_item),
        }

    except HTTPException:
        raise
    except (RuntimeError, OSError, TypeError, ValueError) as e:
        logger.exception("Error refreshing SQL qty for %s", _safe_log_value(barcode))
        raise HTTPException(status_code=500, detail=f"Refresh failed: {str(e)}")


def _calculate_variance(request: VerificationRequest, system_qty: float) -> Optional[float]:
    """Calculates the variance based on verified and damaged quantities."""
    if request.verified_qty is not None:
        total_assets = request.verified_qty + (request.damaged_qty or 0.0)
        return total_assets - system_qty
    return None


def _build_item_update_doc(
    request: VerificationRequest, current_user: dict, existing_item: dict
) -> dict[str, Any]:
    """Builds the update document for the erp_items collection."""
    update_fields = {
        "verified": request.verified,
        "verified_by": current_user["username"],
        "verified_at": datetime.now(timezone.utc).replace(tzinfo=None),
        "last_scanned_at": datetime.now(timezone.utc).replace(tzinfo=None),
    }

    if request.verified_qty is not None:
        update_fields["verified_qty"] = request.verified_qty
        update_fields["variance"] = _calculate_variance(
            request, existing_item.get("stock_qty", 0.0)
        )

    # Map request fields to update_fields, handling None values
    field_mapping = {
        "damaged_qty": "damaged_qty",
        "non_returnable_damaged_qty": "non_returnable_damaged_qty",
        "item_condition": "item_condition",
        "serial_number": "serial_number",
        "notes": "verification_notes",
        "session_id": "session_id",
    }

    for req_field, doc_field in field_mapping.items():
        req_value = getattr(request, req_field)
        if req_value is not None:  # Check for None explicitly, as 0.0 or False are valid
            update_fields[doc_field] = req_value

    # Store observed floor/rack without mutating master location fields.
    if request.floor:
        update_fields["verified_floor"] = request.floor
    if request.rack:
        update_fields["verified_rack"] = request.rack

    return {"$set": update_fields}


def _build_verification_log_doc(
    request: VerificationRequest,
    current_user: dict,
    item: dict,
    variance: Optional[float],
    is_serialized_from_update: Optional[bool],
) -> dict[str, Any]:
    """Builds the document for verification_logs and item_variances collections."""
    return {
        "item_code": item.get("item_code", ""),
        "item_name": item.get("item_name", ""),
        "system_qty": item.get("stock_qty", 0.0),
        "verified_qty": request.verified_qty,
        "damaged_qty": request.damaged_qty,
        "non_returnable_damaged_qty": request.non_returnable_damaged_qty,
        "variance": variance,
        "verified_by": current_user["username"],
        "verified_at": datetime.now(timezone.utc).replace(tzinfo=None),
        "category": item.get("category", ""),
        "subcategory": item.get("subcategory", ""),
        "floor": request.floor or item.get("floor", ""),
        "rack": request.rack or item.get("rack", ""),
        "warehouse": item.get("warehouse", ""),
        "session_id": request.session_id,
        "count_line_id": request.count_line_id,
        "item_condition": request.item_condition,
        "serial_number": request.serial_number,
        "is_serialized": is_serialized_from_update,
    }


@verification_router.patch("/{barcode}/verify")
async def verify_item(
    barcode: str,
    request: VerificationRequest,
    current_user: dict = Depends(get_current_user),
):
    """
    Mark an item as verified/unverified with user tracking and timestamp
    """
    try:
        if request.is_serialized is not None:
            raise HTTPException(
                status_code=400,
                detail=(
                    "CRITICAL: is_serialized is master-data controlled and cannot be set by users."
                ),
            )

        item = await _fetch_item_with_optional_sql_refresh(barcode)
        actual_barcode, actual_item_code, base_filter = _resolve_item_identity(item, barcode)
        expected_stock_qty = item.get("stock_qty")
        update_filter = _build_verification_filter(
            actual_barcode=actual_barcode,
            actual_item_code=actual_item_code,
            requested_barcode=barcode,
            expected_stock_qty=expected_stock_qty,
        )
        if _verification_has_material_conflict(item, request):
            return await _create_conflict_fork_response(
                item=item, request=request, current_user=current_user, barcode=barcode
            )

        variance = _calculate_variance(request, item.get("stock_qty", 0.0))
        update_doc = _build_item_update_doc(request, current_user, item)
        with write_authority("ItemVerificationAPI"):
            result = await _item_service().update_erp_item(update_filter, update_doc)
        if result.matched_count == 0:
            logger.warning(
                "Optimistic Lock Failed for %s. Expected qty: %s",
                _safe_log_value(barcode),
                expected_stock_qty,
            )
            raise HTTPException(
                status_code=409,
                detail=(
                    "Data changed during verification (Optimistic Lock). "
                    "Please refresh and try again."
                ),
            )

        await _invalidate_item_cache(
            actual_barcode=actual_barcode,
            actual_item_code=actual_item_code,
            clear_search_cache=False,
        )
        is_serialized_from_update = update_doc["$set"].get("is_serialized")
        verification_log = _build_verification_log_doc(
            request, current_user, item, variance, is_serialized_from_update
        )
        await _item_service().insert_verification_log(verification_log)
        if variance is not None and variance != 0:
            await _item_service().insert_item_variance(verification_log)

        updated_item = await _fetch_updated_item(update_filter, actual_barcode)
        return {
            "success": True,
            "item": updated_item,
            "variance": variance,
            "message": f"Item {actual_item_code} marked as {'verified' if request.verified else 'unverified'}",
        }

    except HTTPException:
        raise
    except ITEM_VERIFICATION_ERRORS as e:
        logger.error(
            "Error verifying item %s: %s",
            _safe_log_value(barcode),
            _safe_log_value(e, max_length=200),
        )
        raise HTTPException(status_code=500, detail=f"Verification failed: {str(e)}")


@verification_router.get("/filtered")
async def get_filtered_items(
    category: Optional[str] = Query(None, description="Filter by category"),
    subcategory: Optional[str] = Query(None, description="Filter by subcategory"),
    floor: Optional[str] = Query(None, description="Filter by floor"),
    rack: Optional[str] = Query(None, description="Filter by rack"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse"),
    uom_code: Optional[str] = Query(None, description="Filter by UOM code"),
    verified: Optional[bool] = Query(None, description="Filter by verification status"),
    search: Optional[str] = Query(None, description="Search in item name/code"),
    limit: int = Query(100, ge=1, le=1000, description="Maximum results"),
    skip: int = Query(0, ge=0, description="Skip results"),
    current_user: dict = Depends(get_current_user),
):
    """
    Get filtered list of items with various filter options
    """
    try:
        filter_query = build_item_filter_query(
            category=category,
            subcategory=subcategory,
            floor=floor,
            rack=rack,
            warehouse=warehouse,
            uom_code=uom_code,
            verified=verified,
            search=search,
        )

        verified_filter = deepcopy(filter_query)
        verified_filter["verified"] = True

        items, total_count, verified_count = await _item_service().get_filtered_items(
            filter_query=filter_query,
            verified_filter=verified_filter,
            skip=skip,
            limit=limit,
        )

        items = [serialize_item_document(item) for item in items]
        total_qty = sum(item.get("stock_qty", 0.0) for item in items)

        return {
            "success": True,
            "items": items,
            "pagination": {
                "total": total_count,
                "limit": limit,
                "skip": skip,
                "returned": len(items),
            },
            "statistics": {
                "total_items": total_count,
                "verified_items": verified_count,
                "unverified_items": total_count - verified_count,
                "total_qty": total_qty,
            },
        }

    except ITEM_VERIFICATION_ERRORS as e:
        logger.error("Error getting filtered items: %s", _safe_log_value(e, max_length=200))
        raise HTTPException(status_code=500, detail=f"Failed to get items: {str(e)}")


@verification_router.get("/sync")
async def sync_items_for_offline_cache(
    since: Optional[datetime] = Query(
        None, description="Return items updated after this timestamp (ISO 8601)"
    ),
    limit: int = Query(5000, ge=1, le=20000, description="Maximum items to return"),
    current_user: dict = Depends(get_current_user),
):
    """
    Incremental item sync for offline caching.

    Used by the mobile app to keep its local SQLite item table fresh for offline search.
    """
    try:
        query: dict[str, Any] = {"barcode": {"$exists": True, "$ne": ""}}

        if since:
            # Items can be updated via verification, scanning, or master updates.
            query["$or"] = [
                {"last_scanned_at": {"$gt": since}},
                {"verified_at": {"$gt": since}},
                {"last_updated_at": {"$gt": since}},
            ]

        projection = {
            "_id": 0,
            "barcode": 1,
            "item_code": 1,
            "item_name": 1,
            "category": 1,
            "verified": 1,
            "verified_at": 1,
            "last_scanned_at": 1,
        }

        items = await _item_service().sync_items(
            query=query,
            projection=projection,
            limit=limit,
        )
        items = [serialize_item_document(item) for item in items]

        now = datetime.now(timezone.utc).replace(tzinfo=None)
        return {
            "success": True,
            "items": items,
            "count": len(items),
            "since": serialize_mongo_datetime(since) if since else None,
            "server_time": now.isoformat(),
        }

    except ITEM_VERIFICATION_ERRORS as e:
        logger.error("Error syncing items: %s", _safe_log_value(e, max_length=200))
        raise HTTPException(status_code=500, detail=f"Failed to sync items: {str(e)}")


@verification_router.get("/export/csv")
async def export_items_csv(
    category: Optional[str] = Query(None),
    subcategory: Optional[str] = Query(None),
    floor: Optional[str] = Query(None),
    rack: Optional[str] = Query(None),
    warehouse: Optional[str] = Query(None),
    verified: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    max_rows: int = Query(10000, ge=1, le=100000, description="Maximum rows to export"),
    current_user: dict = Depends(get_current_user),
):
    """
    Export filtered items to ERPNext-compatible CSV
    """
    try:
        filter_query = build_item_filter_query(
            category=category,
            subcategory=subcategory,
            floor=floor,
            rack=rack,
            warehouse=warehouse,
            verified=verified,
            search=search,
        )

        async def generate_csv_rows():
            output = io.StringIO()
            writer = csv.DictWriter(
                output, fieldnames=ITEM_EXPORT_FIELDNAMES, extrasaction="ignore"
            )
            writer.writeheader()
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

            async for item in _item_service().iter_item_export_rows(
                filter_query=filter_query,
                max_rows=max_rows,
            ):
                row = _build_erpnext_item_export_row(item)
                writer.writerow(row)
                yield output.getvalue()
                output.seek(0)
                output.truncate(0)

        filename = (
            f"items_erpnext_import_"
            f"{datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.csv"
        )

        return StreamingResponse(
            generate_csv_rows(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except ITEM_VERIFICATION_ERRORS as e:
        logger.error("Error exporting items to CSV: %s", _safe_log_value(e, max_length=200))
        raise HTTPException(status_code=500, detail=f"CSV export failed: {str(e)}")


@verification_router.get("/export/json")
async def export_items_json(
    category: Optional[str] = Query(None),
    subcategory: Optional[str] = Query(None),
    floor: Optional[str] = Query(None),
    rack: Optional[str] = Query(None),
    warehouse: Optional[str] = Query(None),
    verified: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    max_rows: int = Query(10000, ge=1, le=100000, description="Maximum rows to export"),
    current_user: dict = Depends(get_current_user),
):
    """Export filtered items as JSON."""
    try:
        filter_query = build_item_filter_query(
            category=category,
            subcategory=subcategory,
            floor=floor,
            rack=rack,
            warehouse=warehouse,
            verified=verified,
            search=search,
        )

        items = await _item_service().fetch_item_export_rows(
            filter_query=filter_query,
            max_rows=max_rows,
        )
        rows = [_build_erpnext_item_export_row(item) for item in items]

        filename = (
            f"items_export_"
            f"{datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.json"
        )

        return StreamingResponse(
            iter([json.dumps({"items": rows}, default=str)]),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except ITEM_VERIFICATION_ERRORS as e:
        logger.error("Error exporting items to JSON: %s", _safe_log_value(e, max_length=200))
        raise HTTPException(status_code=500, detail=f"JSON export failed: {str(e)}")


@verification_router.get("/export/xlsx")
async def export_items_xlsx(
    category: Optional[str] = Query(None),
    subcategory: Optional[str] = Query(None),
    floor: Optional[str] = Query(None),
    rack: Optional[str] = Query(None),
    warehouse: Optional[str] = Query(None),
    verified: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    max_rows: int = Query(10000, ge=1, le=100000, description="Maximum rows to export"),
    current_user: dict = Depends(get_current_user),
):
    """Export filtered items to ERPNext-compatible Excel."""
    try:
        filter_query = build_item_filter_query(
            category=category,
            subcategory=subcategory,
            floor=floor,
            rack=rack,
            warehouse=warehouse,
            verified=verified,
            search=search,
        )

        items = await _item_service().fetch_item_export_rows(
            filter_query=filter_query,
            max_rows=max_rows,
        )
        rows = [_build_erpnext_item_export_row(item) for item in items]
        content = _render_xlsx_bytes(ITEM_EXPORT_FIELDNAMES, rows)
        filename = (
            f"items_erpnext_import_"
            f"{datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        return StreamingResponse(
            iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except ITEM_VERIFICATION_ERRORS as e:
        logger.error("Error exporting items to XLSX: %s", _safe_log_value(e, max_length=200))
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")


@verification_router.get("/variances")
async def get_variances(
    category: Optional[str] = Query(None),
    floor: Optional[str] = Query(None),
    rack: Optional[str] = Query(None),
    warehouse: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=1000),
    skip: int = Query(0, ge=0),
    current_user: dict = Depends(get_current_user),
):
    """
    Get list of items with variances (verified qty != system qty)
    """
    try:
        filter_query: dict[str, Any] = {}

        if category:
            filter_query["category"] = {"$regex": category, "$options": "i"}

        if floor:
            filter_query["floor"] = {"$regex": floor, "$options": "i"}

        if rack:
            filter_query["rack"] = {"$regex": rack, "$options": "i"}

        if warehouse:
            filter_query["warehouse"] = {"$regex": warehouse, "$options": "i"}

        # Only get variances (non-zero)
        filter_query["variance"] = {"$ne": 0}
        # Optionally, filter for only approved/active count lines if required by logic
        # filter_query["status"] = {"$in": ["approved", "pending"]}

        variances, total_count = await _item_service().get_variances(
            filter_query=filter_query,
            skip=skip,
            limit=limit,
        )

        # Convert ObjectId to string
        for variance in variances:
            variance["_id"] = str(variance["_id"])
            if "counted_at" in variance and isinstance(variance.get("counted_at"), datetime):
                variance["counted_at"] = variance["counted_at"].isoformat()
            # Compatibility map: frontend expects verified_at
            if "verified_at" in variance and isinstance(variance.get("verified_at"), datetime):
                variance["verified_at"] = variance["verified_at"].isoformat()
            elif "counted_at" in variance:
                variance["verified_at"] = variance["counted_at"]

        return {
            "success": True,
            "variances": variances,
            "pagination": {
                "total": total_count,
                "limit": limit,
                "skip": skip,
                "returned": len(variances),
            },
        }

    except ITEM_VERIFICATION_ERRORS as e:
        logger.error("Error getting variances: %s", _safe_log_value(e, max_length=200))
        raise HTTPException(status_code=500, detail=f"Failed to get variances: {str(e)}")


async def _fetch_variance_export_rows(
    *,
    category: Optional[str],
    floor: Optional[str],
    rack: Optional[str],
    warehouse: Optional[str],
    max_rows: int,
) -> list[dict[str, Any]]:
    filter_query: dict[str, Any] = {"variance": {"$ne": 0}}

    if category:
        filter_query["category"] = {"$regex": category, "$options": "i"}
    if floor:
        filter_query["floor"] = {"$regex": floor, "$options": "i"}
    if rack:
        filter_query["rack"] = {"$regex": rack, "$options": "i"}
    if warehouse:
        filter_query["warehouse"] = {"$regex": warehouse, "$options": "i"}

    variances = await _item_service().fetch_variance_export_rows(
        filter_query=filter_query,
        max_rows=max_rows,
    )
    return [_build_erpnext_variance_export_row(variance) for variance in variances]


@verification_router.get("/variances/export/csv")
async def export_variances_csv(
    category: Optional[str] = Query(None),
    floor: Optional[str] = Query(None),
    rack: Optional[str] = Query(None),
    warehouse: Optional[str] = Query(None),
    max_rows: int = Query(10000, ge=1, le=100000, description="Maximum rows to export"),
    current_user: dict = Depends(get_current_user),
):
    """Export variances to ERPNext-compatible CSV."""
    try:
        rows = await _fetch_variance_export_rows(
            category=category,
            floor=floor,
            rack=rack,
            warehouse=warehouse,
            max_rows=max_rows,
        )

        output = io.StringIO()
        writer = csv.DictWriter(
            output, fieldnames=VARIANCE_EXPORT_FIELDNAMES, extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(rows)

        filename = (
            f"variances_erpnext_import_"
            f"{datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.csv"
        )
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ITEM_VERIFICATION_ERRORS as e:
        logger.error(
            "Error exporting variances to CSV: %s",
            _safe_log_value(e, max_length=200),
        )
        raise HTTPException(status_code=500, detail=f"Variance CSV export failed: {str(e)}")


@verification_router.get("/variances/export/xlsx")
async def export_variances_xlsx(
    category: Optional[str] = Query(None),
    floor: Optional[str] = Query(None),
    rack: Optional[str] = Query(None),
    warehouse: Optional[str] = Query(None),
    max_rows: int = Query(10000, ge=1, le=100000, description="Maximum rows to export"),
    current_user: dict = Depends(get_current_user),
):
    """Export variances to ERPNext-compatible Excel."""
    try:
        rows = await _fetch_variance_export_rows(
            category=category,
            floor=floor,
            rack=rack,
            warehouse=warehouse,
            max_rows=max_rows,
        )
        content = _render_xlsx_bytes(VARIANCE_EXPORT_FIELDNAMES, rows)
        filename = (
            f"variances_erpnext_import_"
            f"{datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return StreamingResponse(
            iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except ITEM_VERIFICATION_ERRORS as e:
        logger.error(
            "Error exporting variances to XLSX: %s",
            _safe_log_value(e, max_length=200),
        )
        raise HTTPException(status_code=500, detail=f"Variance Excel export failed: {str(e)}")


@verification_router.get("/live/users")
async def get_live_users(current_user: dict = Depends(get_current_user)):
    """
    Get list of currently active users (users who have verified items in last hour)
    """
    try:
        users = await _item_service().get_live_users()

        result = [
            {
                "username": user["_id"],
                "last_activity": (
                    user["last_activity"].isoformat()
                    if isinstance(user["last_activity"], datetime)
                    else user["last_activity"]
                ),
                "items_verified": user["items_verified"],
            }
            for user in users
        ]

        return {"success": True, "users": result, "count": len(result)}

    except ITEM_VERIFICATION_ERRORS as e:
        logger.error("Error getting live users: %s", _safe_log_value(e, max_length=200))
        raise HTTPException(status_code=500, detail=f"Failed to get live users: {str(e)}")


@verification_router.get("/live/verifications")
async def get_live_verifications(
    limit: int = Query(50, ge=1, le=200),
    current_user: dict = Depends(get_current_user),
):
    """
    Get live feed of recent item verifications
    """
    try:
        verifications = await _item_service().get_live_verifications(limit=limit)

        result = []
        for item in verifications:
            result.append(
                {
                    "item_code": item.get("item_code", ""),
                    "item_name": item.get("item_name", ""),
                    "verified_by": item.get("verified_by", ""),
                    "verified_at": (
                        item.get("verified_at").isoformat() if item.get("verified_at") else None
                    ),
                    "floor": item.get("floor", ""),
                    "rack": item.get("rack", ""),
                    "category": item.get("category", ""),
                    "variance": item.get("variance", 0.0),
                }
            )

        return {"success": True, "verifications": result, "count": len(result)}

    except ITEM_VERIFICATION_ERRORS as e:
        logger.error(
            "Error getting live verifications: %s",
            _safe_log_value(e, max_length=200),
        )
        raise HTTPException(status_code=500, detail=f"Failed to get live verifications: {str(e)}")
